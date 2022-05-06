from __future__ import unicode_literals, print_function

import csv
import json
import logging
import os
import re
import time
import uuid
import hashlib
import copy
from collections import defaultdict
from datetime import timedelta, date
from tempfile import NamedTemporaryFile
from filtration import Expression

import requests
from celery import shared_task, chord, current_task
from celery.exceptions import TimeoutError, MaxRetriesExceededError
from dateutil.rrule import rrule, DAILY
from django.conf import settings
from django.contrib.contenttypes.models import ContentType
from django.core.exceptions import ObjectDoesNotExist
from django.core.mail import send_mail as django_send_mail, send_mass_mail as django_send_mass_mail, \
    mail_admins as django_mail_admins, mail_managers as django_mail_managers
from django.db import transaction
from django.db.models import Q
from django.utils import timezone
from django.utils.timezone import now
from django.utils.translation import ugettext
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string


from poms.accounts.models import Account
from poms.audit.models import AuthLogEntry
from poms.celery_tasks.models import CeleryTask
from poms.common import formula
from poms.common.crypto.AESCipher import AESCipher
from poms.common.crypto.RSACipher import RSACipher
from poms.common.formula import ExpressionEvalError
from poms.common.utils import date_now, isclose
from poms.counterparties.models import Counterparty, Responsible
from poms.currencies.models import Currency, CurrencyHistory
from poms.file_reports.models import FileReport
from poms.instruments.models import Instrument, DailyPricingModel, PricingPolicy, PriceHistory, InstrumentType, \
    PaymentSizeDetail, Periodicity, AccrualCalculationModel
from poms.integrations.models import Task, PriceDownloadScheme, InstrumentDownloadScheme, PricingAutomatedSchedule, \
    AccountMapping, CurrencyMapping, PortfolioMapping, CounterpartyMapping, InstrumentTypeMapping, ResponsibleMapping, \
    Strategy1Mapping, Strategy2Mapping, Strategy3Mapping, DailyPricingModelMapping, PaymentSizeDetailMapping, \
    PriceDownloadSchemeMapping, InstrumentMapping, PeriodicityMapping, AccrualCalculationModelMapping, \
    BloombergDataProviderCredential, ComplexTransactionImportScheme, TransactionFileResult
from poms.integrations.providers.base import get_provider, parse_date_iso, fill_instrument_price, fill_currency_price, \
    AbstractProvider

from poms.integrations.storage import import_file_storage
from poms.portfolios.models import Portfolio
from poms.reports.builders.balance_item import Report, ReportItem
from poms.reports.builders.balance_pl import ReportBuilder
from poms.strategies.models import Strategy1, Strategy2, Strategy3
from poms.system_messages.handlers import send_system_message
from poms.transactions.handlers import TransactionTypeProcess
from poms.transactions.models import EventClass
from poms.users.models import MasterUser, EcosystemDefault
from io import BytesIO

from poms.common.utils import date_now, datetime_now

from .models import ImportConfig
from ..common.jwt import encode_with_jwt
from ..common.websockets import send_websocket_message

import traceback

from ..counterparties.serializers import CounterpartySerializer
from ..csv_import.tasks import handler_instrument_object
from ..obj_attrs.models import GenericAttributeType

_l = logging.getLogger('poms.integrations')

from storages.backends.sftpstorage import SFTPStorage

SFS = SFTPStorage()


@shared_task(name='integrations.health_check')
def health_check_async():
    return True


def health_check():
    result = health_check_async.apply_async()
    try:
        return result.get(timeout=0.5, interval=0.1)
    except TimeoutError:
        pass
    return False


@shared_task(name='integrations.send_mail_async', ignore_result=True)
def send_mail_async(subject, message, from_email, recipient_list, html_message=None):
    django_send_mail(subject, message, from_email, recipient_list, fail_silently=True, html_message=html_message)


def send_mail(subject, message, from_email, recipient_list, html_message=None):
    send_mail_async.apply_async(kwargs={
        'subject': subject,
        'message': message,
        'from_email': from_email,
        'recipient_list': recipient_list,
        'html_message': html_message,
    })


@shared_task(name='integrations.send_mass_mail', ignore_result=True)
def send_mass_mail_async(messages):
    django_send_mass_mail(messages, fail_silently=True)


def send_mass_mail(messages):
    send_mass_mail_async.apply_async(kwargs={
        'messages': messages,
    })


@shared_task(name='integrations.mail_admins', ignore_result=True)
def mail_admins_async(subject, message):
    django_mail_admins(subject, message, fail_silently=True, )


def mail_admins(subject, message):
    mail_admins_async.apply_async(kwargs={
        'subject': subject,
        'message': message,
    })


@shared_task(name='integrations.mail_managers', ignore_result=True)
def mail_managers_async(subject, message):
    django_mail_managers(subject, message, fail_silently=True, )


def mail_managers(subject, message):
    mail_managers_async.apply_async(kwargs={
        'subject': subject,
        'message': message,
    })


@shared_task(name='integrations.auth_log_statistics', ignore_result=True)
def auth_log_statistics():
    logged_in_count = AuthLogEntry.objects.filter(is_success=True).count()
    login_failed_count = AuthLogEntry.objects.filter(is_success=False).count()
    _l.debug('auth (total): logged_in=%s, login_failed=%s', logged_in_count, login_failed_count)

    now = timezone.now().date()
    logged_in_count = AuthLogEntry.objects.filter(is_success=True, date__startswith=now).count()
    login_failed_count = AuthLogEntry.objects.filter(is_success=False, date__startswith=now).count()
    _l.debug('auth (today): logged_in=%s, login_failed=%s', logged_in_count, login_failed_count)


@shared_task(name='integrations.download_instrument', bind=True, ignore_result=False)
def download_instrument_async(self, task_id=None):
    task = Task.objects.get(pk=task_id)
    _l.debug('download_instrument_async: master_user_id=%s, task=%s', task.master_user_id, task.info)

    task.add_celery_task_id(self.request.id)

    try:
        provider = get_provider(task.master_user, task.provider_id)
    except Exception:
        _l.debug('provider load error', exc_info=True)
        task.status = Task.STATUS_ERROR
        task.save()
        raise

    if provider is None:
        _l.debug('provider not found')
        task.status = Task.STATUS_ERROR
        task.save()
        return

    if task.status not in [Task.STATUS_PENDING, Task.STATUS_WAIT_RESPONSE]:
        _l.debug('invalid task status')
        return
    options = task.options_object

    try:
        result, is_ready = provider.download_instrument(options)
    except Exception:
        _l.error('provider processing error', exc_info=True)
        task.status = Task.STATUS_ERROR
    else:
        if is_ready:
            task.status = Task.STATUS_DONE
            task.result_object = result
        else:
            task.status = Task.STATUS_WAIT_RESPONSE

    response_id = options.get('response_id', None)
    if response_id:
        task.response_id = response_id

    task.options_object = options
    task.save()

    if task.status == Task.STATUS_WAIT_RESPONSE:
        if self.request.is_eager:
            time.sleep(provider.get_retry_delay())
        try:
            self.retry(countdown=provider.get_retry_delay(), max_retries=provider.get_max_retries())
            # self.retry(countdown=provider.get_retry_delay(), max_retries=provider.get_max_retries(), throw=False)
        except MaxRetriesExceededError:
            task.status = Task.STATUS_TIMEOUT
            task.save()
        return

    return task_id


def download_instrument(instrument_code=None, instrument_download_scheme=None, master_user=None, member=None,
                        task=None, value_overrides=None):
    _l.debug('download_pricing: master_user_id=%s, task=%s, instrument_code=%s, instrument_download_scheme=%s',
            getattr(master_user, 'id', None), getattr(task, 'info', None), instrument_code, instrument_download_scheme)

    if task is None:
        provider = get_provider(instrument_download_scheme.master_user, instrument_download_scheme.provider)
        if not provider.is_valid_reference(instrument_code):
            raise ValueError('Invalid instrument_code value')

        options = {
            'instrument_download_scheme_id': instrument_download_scheme.id,
            'instrument_code': instrument_code,
        }
        with transaction.atomic():
            task = Task(
                master_user=master_user,
                member=member,
                provider=instrument_download_scheme.provider,
                status=Task.STATUS_PENDING,
                action=Task.ACTION_INSTRUMENT
            )
            task.options_object = options
            task.save()
            # transaction.on_commit(
            #     lambda: download_instrument_async.apply_async(kwargs={'task_id': task.id}, countdown=1))
            transaction.on_commit(lambda: download_instrument_async.apply_async(kwargs={'task_id': task.id}))
        return task, None, None
    else:
        if task.status == Task.STATUS_DONE:
            provider = get_provider(task.master_user, task.provider_id)

            options = task.options_object
            values = task.result_object.copy()
            if value_overrides:
                values.update(value_overrides)

            instrument_download_scheme_id = options['instrument_download_scheme_id']
            instrument_download_scheme = InstrumentDownloadScheme.objects.get(pk=instrument_download_scheme_id)

            instrument, errors = provider.create_instrument(instrument_download_scheme, values)
            return task, instrument, errors
        return task, None, None

class ProxyUser(object):

    def __init__(self, member, master_user):
        self.member = member
        self.master_user = master_user


class ProxyRequest(object):

    def __init__(self, user):
        self.user = user



def create_instrument_cbond(data, master_user, member):

    try:

        from poms.instruments.serializers import InstrumentSerializer

        ecosystem_defaults = EcosystemDefault.objects.get(master_user=master_user)
        content_type = ContentType.objects.get(model="instrument", app_label="instruments")

        proxy_user = ProxyUser(member, master_user)
        proxy_request = ProxyRequest(proxy_user)

        context = {'master_user': master_user,
                   'request': proxy_request}

        instrument_data = {}

        for key, value in data.items():

            if key == 'attributes':

                for attr_key, attr_value in data['attributes'].items():

                    if attr_value == 'null':
                        instrument_data[attr_key] = None
                    else:
                        instrument_data[attr_key] = attr_value

            else:

                if value == 'null':
                    instrument_data[key] = None
                else:
                    instrument_data[key] = value



        attribute_types =  GenericAttributeType.objects.filter(master_user=master_user,
                                                               content_type=content_type)

        instrument_type = None

        try:

            instrument_type = InstrumentType.objects.get(master_user=master_user,
                                                         user_code=instrument_data['instrument_type'])

        except Exception as e:

            _l.info('Instrument Type %s is not found %s' % (instrument_data['instrument_type'], e))

            raise Exception("Instrument Type %s is not found %s" % (instrument_data['instrument_type'], e))

        object_data = handler_instrument_object(instrument_data, instrument_type, master_user, ecosystem_defaults, attribute_types)


        try:

            instance = Instrument.objects.get(master_user=master_user, user_code=object_data['user_code'])

            serializer = InstrumentSerializer(data=object_data, context=context, instance=instance)
        except Instrument.DoesNotExist:

            serializer = InstrumentSerializer(data=object_data, context=context)

        is_valid = serializer.is_valid()

        if is_valid:
            instrument = serializer.save()

            _l.info("Instrument is imported successfully")

            return instrument
        else:
            _l.info('InstrumentExternalAPIViewSet error %s' % serializer.errors)
            raise Exception(serializer.errors)

    except Exception as e:
        _l.info('InstrumentExternalAPIViewSet error %s' % e)
        _l.info(traceback.format_exc())
        raise Exception(e)


def create_currency_cbond(data, master_user, member):

    try:

        from poms.currencies.serializers import CurrencySerializer

        ecosystem_defaults = EcosystemDefault.objects.get(master_user=master_user)
        content_type = ContentType.objects.get(model="currency", app_label="currencies")

        proxy_user = ProxyUser(member, master_user)
        proxy_request = ProxyRequest(proxy_user)

        context = {'master_user': master_user,
                   'request': proxy_request}

        currency_data = {}

        for key, value in data.items():

            if key == 'attributes':

                for attr_key, attr_value in data['attributes'].items():

                    if attr_value == 'null':
                        currency_data[attr_key] = None
                    else:
                        currency_data[attr_key] = attr_value

            else:

                if value == 'null':
                    currency_data[key] = None
                else:
                    currency_data[key] = value



        attribute_types =  GenericAttributeType.objects.filter(master_user=master_user,
                                                               content_type=content_type)

        try:

            instance = Currency.objects.get(master_user=master_user, user_code=currency_data['user_code'])

            serializer = CurrencySerializer(data=currency_data, context=context, instance=instance)
        except Currency.DoesNotExist:

            serializer = CurrencySerializer(data=currency_data, context=context)

        is_valid = serializer.is_valid()

        if is_valid:
            currency = serializer.save()

            _l.info("Currency is imported successfully")

            return currency
        else:
            _l.info('CurrencyExternalAPIViewSet error %s' % serializer.errors)
            raise Exception(serializer.errors)

    except Exception as e:
        _l.info('CurrencyExternalAPIViewSet error %s' % e)
        _l.info(traceback.format_exc())
        raise Exception(e)


def download_instrument_cbond(instrument_code=None, master_user=None, member=None):

    errors = []

    try:
        _l.debug('download_instrument_cbond: master_user_id=%s, instrument_code=%s',
                 getattr(master_user, 'id', None), instrument_code)

        options = {

            'isin': instrument_code,
        }
        with transaction.atomic():
            task = Task(
                master_user=master_user,
                member=member,
                status=Task.STATUS_PENDING,
                action=Task.ACTION_INSTRUMENT
            )
            task.options_object = options
            task.save()

            headers = {'Content-type': 'application/json'}

            payload_jwt = {
                "sub":  settings.BASE_API_URL, #"user_id_or_name",
                "role": 0 # 0 -- ordinary user, 1 -- admin (access to /loadfi and /loadeq)
            }

            token = encode_with_jwt(payload_jwt)

            headers['Authorization'] = 'Bearer %s' % token

            options['request_id'] = task.pk
            options['base_api_url'] = settings.BASE_API_URL
            options['token'] = 'fd09a190279e45a2bbb52fcabb7899bd'

            options['data'] = {}


            response = None

            # OLD ASYNC CODE
            # try:
            #     response = requests.post(url=str(settings.CBONDS_BROKER_URL) + '/request-instrument/', data=json.dumps(options), headers=headers)
            #     _l.info('response download_instrument_cbond %s' % response)
            # except Exception as e:
            #     _l.debug("Can't send request to CBONDS BROKER. %s" % e)

            _l.info('options %s' % options)
            _l.info('settings.CBONDS_BROKER_URL %s' % settings.CBONDS_BROKER_URL)

            try:
                response = requests.post(url=str(settings.CBONDS_BROKER_URL) + 'export/', data=json.dumps(options), headers=headers)
                _l.info('response download_instrument_cbond %s' % response)
                _l.info('data response.text %s ' % response.text)
            except Exception as e:
                _l.debug("Can't send request to CBONDS BROKER. %s" % e)

                errors.append('Request to broker failed. %s' % str(e))

            try:
                data = response.json()
            except Exception as e:

                errors.append("Could not parse response from broker. %s" % response.text)
                return task, errors
            try:

                if 'items' in data['data']:

                    for item in data['data']['items']:
                        instrument = create_instrument_cbond(item, master_user, member)

                else:

                    instrument = create_instrument_cbond(data['data'], master_user, member)

                result = {
                    "instrument_id": instrument.pk
                }

                task.result_object = result

                task.save()
                return task, errors

            except Exception as e:
                errors.append("Could not create instrument. %s" % str(e))
                return task, errors

            _l.info('data %s ' % data)

            return task, errors

    except Exception as e:
        _l.info("error %s " % e)
        _l.info(traceback.format_exc())

        errors.append('Something went wrong. %s' % str(e))

        return None, errors


def download_currency_cbond(currency_code=None, master_user=None, member=None):

    errors = []

    try:
        _l.debug('download_currency_cbond: master_user_id=%s, currency_code=%s',
                 getattr(master_user, 'id', None), currency_code)

        options = {
            'code': currency_code,
        }
        with transaction.atomic():
            task = Task(
                master_user=master_user,
                member=member,
                status=Task.STATUS_PENDING,
                action=Task.ACTION_INSTRUMENT
            )
            task.options_object = options
            task.save()

            headers = {'Content-type': 'application/json'}

            payload_jwt = {
                "sub":  settings.BASE_API_URL, #"user_id_or_name",
                "role": 0 # 0 -- ordinary user, 1 -- admin (access to /loadfi and /loadeq)
            }

            token = encode_with_jwt(payload_jwt)

            headers['Authorization'] = 'Bearer %s' % token

            options['request_id'] = task.pk
            options['base_api_url'] = settings.BASE_API_URL
            options['token'] = 'fd09a190279e45a2bbb52fcabb7899bd'

            options['data'] = {}


            response = None

            # OLD ASYNC CODE
            # try:
            #     response = requests.post(url=str(settings.CBONDS_BROKER_URL) + '/request-instrument/', data=json.dumps(options), headers=headers)
            #     _l.info('response download_instrument_cbond %s' % response)
            # except Exception as e:
            #     _l.debug("Can't send request to CBONDS BROKER. %s" % e)

            _l.info('options %s' % options)
            _l.info('settings.CBONDS_BROKER_URL %s' % settings.CBONDS_BROKER_URL)

            try:
                # TODO refactor to /export/currency when available
                response = requests.get(url=str(settings.CBONDS_BROKER_URL) + 'instr/currency/' + currency_code, headers=headers)
                _l.info('response download_currency_cbond %s' % response)
                _l.info('data response.text %s ' % response.text)
            except Exception as e:
                _l.debug("Can't send request to CBONDS BROKER. %s" % e)

                errors.append('Request to broker failed. %s' % str(e))

            try:
                data = response.json()
            except Exception as e:

                errors.append("Could not parse response from broker. %s" % response.text)
                return task, errors
            try:

                if 'items' in data['data']:

                    for item in data['data']['items']:
                        currency = create_currency_cbond(item, master_user, member)

                else:

                    currency = create_currency_cbond(data['data'], master_user, member)

                result = {
                    "currency_id": currency.pk
                }

                task.result_object = result

                task.save()
                return task, errors

            except Exception as e:
                errors.append("Could not create currency. %s" % str(e))
                return task, errors

            _l.info('data %s ' % data)

            return task, errors

    except Exception as e:
        _l.info("error %s " % e)
        _l.info(traceback.format_exc())

        errors.append('Something went wrong. %s' % str(e))

        return None, errors


@shared_task(name='integrations.download_instrument_cbond_task', bind=True, ignore_result=False)
def download_instrument_cbond_task(self, task_id):

    task = Task.objects.get(pk=task_id)

    download_instrument_cbond(task.options['user_code'], task.master_user, task.member)


def download_unified_data(id=None, entity_type=None, master_user=None, member=None,
                          task=None, value_overrides=None):

    errors = []

    try:

        with transaction.atomic():
            task = Task(
                master_user=master_user,
                member=member,
                status=Task.STATUS_PENDING,
                action=Task.ACTION_INSTRUMENT
            )
            task.options_object = {
                "entity_type": entity_type,
                "id": id
            }
            task.save()

            headers = {'Content-type': 'application/json'}


            response = None

            path = ''

            if entity_type == 'counterparty':
                path = 'company'

            try:
                response = requests.get(url=str(settings.UNIFIED_DATA_PROVIDER_URL) + 'data/' + path + '/' + id + '/', headers=headers)
                _l.info('response download_unified_data %s' % response)
                _l.info('data response.text %s ' % response.text)
            except Exception as e:
                _l.debug("Can't send request to Unified Data Provider. %s" % e)

                errors.append('Request to unified data provider. %s' % str(e))

            try:
                data = response.json()
            except Exception as e:

                errors.append("Could not parse response from unified data provider. %s" % response.text)
                return task, errors
            try:

                obj_data = data

                proxy_user = ProxyUser(member, master_user)
                proxy_request = ProxyRequest(proxy_user)

                context = {
                    'request': proxy_request
                }

                ecosystem_defaults = EcosystemDefault.objects.get(master_user=master_user)

                record = None

                if entity_type == 'counterparty':

                    obj_data['group'] = ecosystem_defaults.counterparty_group_id

                    serializer = CounterpartySerializer(data=obj_data,
                                                           context=context)
                    serializer.is_valid(raise_exception=True)
                    record = serializer.save()


                result = {
                    "id": record.pk
                }

                task.result_object = result

                task.save()
                return task, errors

            except Exception as e:
                errors.append("Could not create record. %s" % str(e))
                return task, errors

            return task, errors

    except Exception as e:
        _l.info("error %s " % e)
        _l.info(traceback.format_exc())

        errors.append('Something went wrong. %s' % str(e))

        return None, errors

@shared_task(name='integrations.download_instrument_pricing_async', bind=True, ignore_result=False)
def download_instrument_pricing_async(self, task_id):
    task = Task.objects.get(pk=task_id)
    _l.debug('download_instrument_pricing_async: master_user_id=%s, task=%s', task.master_user_id, task.info)

    task.add_celery_task_id(self.request.id)

    try:
        provider = get_provider(task.master_user, task.provider_id)
    except Exception:
        _l.debug('provider load error', exc_info=True)
        task.status = Task.STATUS_ERROR
        task.save()
        return

    if provider is None:
        _l.debug('provider not found')
        task.status = Task.STATUS_ERROR
        task.save()
        return

    if task.status not in [Task.STATUS_PENDING, Task.STATUS_WAIT_RESPONSE]:
        _l.warn('invalid task status')
        return

    options = task.options_object

    try:
        result, is_ready = provider.download_instrument_pricing(options)
    except Exception:
        _l.warn("provider processing error", exc_info=True)
        task.status = Task.STATUS_ERROR
    else:
        if is_ready:
            task.status = Task.STATUS_DONE
            task.result_object = result
        else:
            task.status = Task.STATUS_WAIT_RESPONSE

    response_id = options.get('response_id', None)
    if response_id:
        task.response_id = response_id
    task.options_object = options
    task.save()

    if task.status == Task.STATUS_WAIT_RESPONSE:
        if self.request.is_eager:
            time.sleep(provider.get_retry_delay())
        try:
            self.retry(countdown=provider.get_retry_delay(), max_retries=provider.get_max_retries())
            # self.retry(countdown=provider.get_retry_delay(), max_retries=provider.get_max_retries(), throw=False)
        except MaxRetriesExceededError:
            task.status = Task.STATUS_TIMEOUT
            task.save()
        return

    return task_id


@shared_task(name='integrations.test_certificate_async', bind=True, ignore_result=False)
def test_certificate_async(self, task_id):
    task = Task.objects.get(pk=task_id)
    _l.debug('handle_test_certificate_async: master_user_id=%s, task=%s', task.master_user_id, task.info)

    task.add_celery_task_id(self.request.id)

    try:
        provider = get_provider(task.master_user, task.provider_id)
    except Exception:
        _l.debug('provider load error', exc_info=True)
        task.status = Task.STATUS_ERROR
        task.save()
        return

    if provider is None:
        _l.debug('provider not found')
        task.status = Task.STATUS_ERROR
        task.save()
        return

    if task.status not in [Task.STATUS_PENDING, Task.STATUS_WAIT_RESPONSE]:
        _l.warn('invalid task status')
        return

    options = task.options_object

    try:
        result = provider.test_certificate(options)
    except Exception as e:
        _l.warn("provider processing error", exc_info=True)
        task.status = Task.STATUS_ERROR

        task.save()
        return
    else:
        _l.debug('handle_test_certificate_async task: result %s' % result)
        _l.debug('handle_test_certificate_async task: result is authorized %s' % result['is_authorized'])

        task.status = Task.STATUS_DONE
        task.result_object = result

        task.options_object = options
        task.save()

        import_config = None

        try:

            import_config = BloombergDataProviderCredential.objects.get(master_user=task.master_user)

            _l.debug('handle_test_certificate_async get actual bloomberg credential')

        except (BloombergDataProviderCredential.DoesNotExist, Exception) as e:

            _l.debug('handle_test_certificate_async get config error', e)

            import_config = ImportConfig.objects.get(master_user=task.master_user, provider=1)

        import_config.is_valid = result['is_authorized']
        import_config.save()

        _l.debug('handle_test_certificate_async import_config: import_config id', import_config.id)
        _l.debug('handle_test_certificate_async import_config: import_config=%s, is_valid=%s', import_config,
                import_config.is_valid)
        _l.debug('handle_test_certificate_async task: master_user_id=%s, task=%s', task.master_user_id, task.result)
        _l.debug('handle_test_certificate_async task.status: ', task.status)

    if task.status == Task.STATUS_WAIT_RESPONSE:
        if self.request.is_eager:
            time.sleep(provider.get_retry_delay())
        try:
            self.retry(countdown=provider.get_retry_delay(), max_retries=provider.get_max_retries())
            # self.retry(countdown=provider.get_retry_delay(), max_retries=provider.get_max_retries(), throw=False)
        except MaxRetriesExceededError:
            task.status = Task.STATUS_TIMEOUT
            task.save()
        return

    return task_id


# DEPRECATED SINCE 22.09.2020 DELETE SOON
@shared_task(name='integrations.download_currency_pricing_async', bind=True, ignore_result=False)
def download_currency_pricing_async(self, task_id):
    task = Task.objects.get(pk=task_id)
    _l.debug('download_currency_pricing_async: master_user_id=%s, task=%s', task.master_user_id, task.info)

    task.add_celery_task_id(self.request.id)

    try:
        provider = get_provider(task.master_user, task.provider_id)
    except Exception:
        _l.debug('provider load error', exc_info=True)
        task.status = Task.STATUS_ERROR
        task.save()
        return

    if provider is None:
        _l.debug('provider not found')
        task.status = Task.STATUS_ERROR
        task.save()
        return

    if task.status not in [Task.STATUS_PENDING, Task.STATUS_WAIT_RESPONSE]:
        _l.warn('invalid task status')
        return

    options = task.options_object

    try:
        result, is_ready = provider.download_currency_pricing(options)
    except Exception:
        _l.warn("provider processing error", exc_info=True)
        task.status = Task.STATUS_ERROR
    else:
        if is_ready:
            task.status = Task.STATUS_DONE
            task.result_object = result
        else:
            task.status = Task.STATUS_WAIT_RESPONSE

    response_id = options.get('response_id', None)
    if response_id:
        task.response_id = response_id

    task.options_object = options
    task.save()

    if task.status == Task.STATUS_WAIT_RESPONSE:
        if self.request.is_eager:
            time.sleep(provider.get_retry_delay())
        try:
            self.retry(countdown=provider.get_retry_delay(), max_retries=provider.get_max_retries())
            # self.retry(countdown=provider.get_retry_delay(), max_retries=provider.get_max_retries(), throw=False)
        except MaxRetriesExceededError:
            task.status = Task.STATUS_TIMEOUT
            task.save()
        return

    return task_id


# DEPRECATED SINCE 22.09.2020 DELETE SOON
@shared_task(name='integrations.download_pricing_async', bind=True, ignore_result=False)
def download_pricing_async(self, task_id):
    _l.info('download pricing async')
    task = Task.objects.get(pk=task_id)
    _l.debug('download_pricing_async: master_user_id=%s, task=%s', task.master_user_id, task.info)

    if task.status not in [Task.STATUS_PENDING, Task.STATUS_WAIT_RESPONSE]:
        return

    task.add_celery_task_id(self.request.id)
    task.status = Task.STATUS_WAIT_RESPONSE

    master_user = task.master_user
    options = task.options_object

    instruments = Instrument.objects.select_related('price_download_scheme').filter(
        master_user=master_user
    ).exclude(
        daily_pricing_model=DailyPricingModel.SKIP
    )
    _l.debug('instruments: %s', [i.id for i in instruments])

    currencies = Currency.objects.select_related('price_download_scheme').filter(
        master_user=master_user
    ).exclude(
        daily_pricing_model=DailyPricingModel.SKIP
    )
    _l.debug('currencies: %s', [i.id for i in currencies])

    instruments_always = set()
    instruments_if_open = set()
    instruments_opened = set()
    instruments_default = set()

    currencies_always = set()
    currencies_if_open = set()
    currencies_opened = set()
    currencies_default = set()

    for i in instruments:

        if i.daily_pricing_model_id in [DailyPricingModel.FORMULA_IF_OPEN, DailyPricingModel.PROVIDER_IF_OPEN]:
            instruments_if_open.add(i.id)

        elif i.daily_pricing_model_id in [DailyPricingModel.FORMULA_ALWAYS, DailyPricingModel.PROVIDER_ALWAYS]:
            instruments_always.add(i.id)
            if i.pricing_currency_id:
                currencies_always.add(i.pricing_currency_id)
            if i.accrued_currency_id:
                currencies_always.add(i.accrued_currency_id)

        elif i.daily_pricing_model_id in [DailyPricingModel.DEFAULT]:
            instruments_default.add(i.id)

    for i in currencies:

        if i.daily_pricing_model_id in [DailyPricingModel.FORMULA_IF_OPEN, DailyPricingModel.PROVIDER_IF_OPEN]:
            currencies_if_open.add(i.id)

        elif i.daily_pricing_model_id in [DailyPricingModel.FORMULA_ALWAYS, DailyPricingModel.PROVIDER_ALWAYS]:
            currencies_always.add(i.id)

        elif i.daily_pricing_model_id in [DailyPricingModel.DEFAULT]:
            currencies_default.add(i.id)

    _l.debug('always: instruments=%s, currencies=%s',
            sorted(instruments_always), sorted(currencies_always))

    balance_date = parse_date_iso(options['balance_date'])
    _l.debug('calculate position report on %s for: instruments=%s, currencies=%s',
            balance_date, sorted(instruments_if_open), sorted(currencies_if_open))

    if balance_date and (instruments_if_open or currencies_if_open):
        owner_or_admin = task.master_user.members.filter(Q(is_owner=True) | Q(is_admin=True)).first()
        report = Report(master_user=task.master_user, member=owner_or_admin, report_date=balance_date)
        _l.debug('calculate position report: %s', report)
        builder = ReportBuilder(instance=report)
        builder.build_position_only()
        for i in report.items:
            if i.type == ReportItem.TYPE_INSTRUMENT and not isclose(i.pos_size, 0.0):
                if i.instr:
                    instruments_opened.add(i.instr.id)
                    if i.instr.pricing_currency_id:
                        currencies_opened.add(i.instr.pricing_currency_id)
                    if i.instr.accrued_currency_id:
                        currencies_opened.add(i.instr.accrued_currency_id)
                if i.trn_ccy:
                    currencies_opened.add(i.trn_ccy.id)
            elif i.type == ReportItem.TYPE_CURRENCY and not isclose(i.pos_size, 0.0):
                if i.ccy:
                    currencies_opened.add(i.ccy.id)
                if i.trn_ccy:
                    currencies_opened.add(i.trn_ccy.id)
        _l.debug('opened: instruments=%s, currencies=%s', sorted(instruments_opened), sorted(currencies_opened))

    instruments = instruments.filter(pk__in=(instruments_always | instruments_opened | instruments_default))
    _l.debug('instruments: %s', [i.id for i in instruments])

    currencies = currencies.filter(pk__in=(currencies_always | currencies_opened | currencies_default))
    _l.debug('currencies: %s', [i.id for i in currencies])

    price_download_schemes = {}

    instruments_by_scheme = defaultdict(list)
    instruments_by_formula = []
    instruments_by_default = []

    for i in instruments:
        if i.daily_pricing_model_id in [DailyPricingModel.PROVIDER_ALWAYS, DailyPricingModel.PROVIDER_IF_OPEN]:
            if i.price_download_scheme_id and i.reference_for_pricing:
                instruments_by_scheme[i.price_download_scheme.id].append(i)
                price_download_schemes[i.price_download_scheme.id] = i.price_download_scheme
        elif i.daily_pricing_model_id in [DailyPricingModel.FORMULA_ALWAYS, DailyPricingModel.FORMULA_IF_OPEN]:
            instruments_by_formula.append(i)
        elif i.daily_pricing_model_id in [DailyPricingModel.DEFAULT]:
            instruments_by_default.append(i)

        _l.debug('instruments_by_scheme: %s', instruments_by_scheme)
    _l.debug('instruments_by_formula: %s', instruments_by_formula)
    _l.debug('instruments_by_default: %s', instruments_by_default)

    currencies_by_scheme = defaultdict(list)
    currencies_by_default = []
    for c in currencies:
        if c.daily_pricing_model_id in [DailyPricingModel.PROVIDER_ALWAYS, DailyPricingModel.PROVIDER_IF_OPEN]:
            if c.price_download_scheme_id and c.reference_for_pricing:
                currencies_by_scheme[c.price_download_scheme.id].append(c)
                price_download_schemes[c.price_download_scheme.id] = c.price_download_scheme
        elif c.daily_pricing_model_id in [DailyPricingModel.DEFAULT]:
            currencies_by_default.append(c)

    _l.debug('currencies_by_scheme: %s', currencies_by_scheme)
    _l.debug('currencies_by_default: %s', currencies_by_default)

    # sub_tasks = []
    # celery_sub_tasks = []

    instrument_sub_tasks = []
    currency_sub_tasks = []

    def sub_tasks_submit():
        celery_sub_tasks = []

        for sub_task_id in instrument_sub_tasks:
            ct = download_instrument_pricing_async.s(task_id=sub_task_id)
            celery_sub_tasks.append(ct)

        for sub_task_id in currency_sub_tasks:
            ct = download_currency_pricing_async.s(task_id=sub_task_id)
            celery_sub_tasks.append(ct)

        _l.debug('celery_sub_tasks: %s', celery_sub_tasks)
        if celery_sub_tasks:
            _l.debug('use chord')
            sub_tasks = instrument_sub_tasks + currency_sub_tasks
            chord(celery_sub_tasks, download_pricing_wait.si(sub_tasks_id=sub_tasks, task_id=task_id)).apply_async()
        else:
            _l.debug('use apply_async')
            download_pricing_wait.apply_async(kwargs={'sub_tasks_id': [], 'task_id': task_id})

    with transaction.atomic():
        instrument_task = defaultdict(list)
        for scheme_id, instruments0 in instruments_by_scheme.items():
            price_download_scheme = price_download_schemes[scheme_id]
            sub_options = options.copy()
            sub_options['price_download_scheme_id'] = price_download_scheme.id
            sub_options['instruments'] = [i.reference_for_pricing for i in instruments0]
            sub_options['instruments_pk'] = [i.id for i in instruments0]

            sub_task = Task(
                master_user=master_user,
                member=task.member,
                parent=task,
                provider=price_download_scheme.provider,
                status=Task.STATUS_PENDING,
                action=Task.ACTION_PRICING
            )
            sub_task.options_object = sub_options
            sub_task.save()

            instrument_sub_tasks.append(sub_task.id)
            # celery_sub_task = download_instrument_pricing_async.apply_async(kwargs={'task_id': sub_task.id})
            # celery_sub_tasks.append(celery_sub_task)

            for i in instruments0:
                instrument_task[i.id] = sub_task.id

        # for manual formula& calculate on final stage
        for i in instruments_by_formula:
            instrument_task[i.id] = None

        for i in instruments_by_default:
            instrument_task[i.id] = 'instrument_default'

        currency_task = defaultdict(list)
        for scheme_id, currencies0 in currencies_by_scheme.items():
            price_download_scheme = price_download_schemes[scheme_id]
            sub_options = options.copy()
            sub_options['price_download_scheme_id'] = price_download_scheme.id
            sub_options['currencies'] = [i.reference_for_pricing for i in currencies0]
            sub_options['currencies_pk'] = [i.id for i in currencies0]

            sub_task = Task(
                master_user=master_user,
                member=task.member,
                parent=task,
                provider=price_download_scheme.provider,
                status=Task.STATUS_PENDING,
                action=Task.ACTION_PRICING
            )
            sub_task.options_object = sub_options
            sub_task.save()

            currency_sub_tasks.append(sub_task.id)
            # celery_sub_task = download_currency_pricing_async.apply_async(kwargs={'task_id': sub_task.id})
            # celery_sub_tasks.append(celery_sub_task)

            for i in currencies0:
                currency_task[i.id] = sub_task.id

        for i in currencies_by_default:
            currency_task[i.id] = 'currency_default'

        options['instrument_task'] = instrument_task
        options['currency_task'] = currency_task
        options['sub_tasks'] = instrument_sub_tasks + currency_sub_tasks

        task.options_object = options
        task.save()

        # if self.request.is_eager:
        #     download_pricing_wait.apply_async(kwargs={'sub_tasks_id': sub_tasks, 'task_id': task_id})
        # else:
        #     if celery_sub_tasks:
        #         chord(celery_sub_tasks, download_pricing_wait.s(task_id=task_id)).apply_async()
        #     else:
        #         download_pricing_wait.apply_async(kwargs={'sub_tasks_id': [], 'task_id': task_id})

        transaction.on_commit(sub_tasks_submit)

    return task_id


# DEPRECATED SINCE 22.09.2020 DELETE SOON
# @shared_task(name='integrations.download_pricing_wait', bind=True, ignore_result=False)
# def download_pricing_wait(self, sub_tasks_id, task_id):
#     _l.info('download pricing wait')
#     task = Task.objects.get(pk=task_id)
#     celery_task = CeleryTask.objects.get(task_id=task_id, master_user=task.master_user)
#     _l.debug('download_pricing_wait: master_user_id=%s, task=%s', task.master_user_id, task.info)
#
#     if task.status != Task.STATUS_WAIT_RESPONSE:
#         return
#
#     task.add_celery_task_id(self.request.id)
#
#     pricing_policies = [p for p in PricingPolicy.objects.filter(master_user=task.master_user)]
#
#     options = task.options_object
#     date_from = parse_date_iso(options['date_from'])
#     date_to = parse_date_iso(options['date_to'])
#     is_yesterday = options['is_yesterday']
#     override_existed = options['override_existed']
#     fill_days = options['fill_days']
#     # sub_tasks_id = options['sub_tasks']
#     instrument_task = options['instrument_task']
#     currency_task = options['currency_task']
#
#     result = {}
#     errors = {}
#     instruments_prices = []
#     currencies_prices = []
#
#     _l.debug('instrument_task: %s', instrument_task)
#     _l.debug('currency_task: %s', currency_task)
#
#     instruments_pk = [int(pk) for pk in instrument_task.keys()]
#     _l.debug('instruments_pk: %s', instruments_pk)
#     currencies_pk = [int(pk) for pk in currency_task.keys()]
#     _l.debug('currencies_pk: %s', currencies_pk)
#
#     _l.debug('sub_tasks_id: %s', sub_tasks_id)
#     for sub_task in Task.objects.filter(pk__in=sub_tasks_id):
#         _l.debug('sub_task: %s', sub_task.info)
#         if sub_task.status != Task.STATUS_DONE:
#             continue
#
#         provider = get_provider(task=sub_task)
#
#         sub_task_options = sub_task.options_object
#
#         if 'instruments_pk' in sub_task_options:
#             task_instruments_pk = sub_task_options['instruments_pk']
#             task_instruments = Instrument.objects.filter(pk__in=task_instruments_pk)
#
#             price_download_scheme_id = sub_task_options['price_download_scheme_id']
#             price_download_scheme = PriceDownloadScheme.objects.get(pk=price_download_scheme_id)
#
#             sub_task_instruments_prices, sub_task_errors = provider.create_instrument_pricing(
#                 price_download_scheme=price_download_scheme,
#                 options=sub_task_options,
#                 values=sub_task.result_object,
#                 instruments=task_instruments,
#                 pricing_policies=pricing_policies
#             )
#
#             instruments_prices += sub_task_instruments_prices
#             errors.update(sub_task_errors)
#
#         elif 'currencies_pk' in sub_task_options:
#             task_currencies_pk = sub_task_options['currencies_pk']
#             task_currencies = Currency.objects.filter(pk__in=task_currencies_pk)
#
#             price_download_scheme_id = sub_task_options['price_download_scheme_id']
#             price_download_scheme = PriceDownloadScheme.objects.get(pk=price_download_scheme_id)
#
#             sub_task_currencies_prices, sub_task_errors = provider.create_currency_pricing(
#                 price_download_scheme=price_download_scheme,
#                 options=sub_task_options,
#                 values=sub_task.result_object,
#                 currencies=task_currencies,
#                 pricing_policies=pricing_policies
#             )
#
#             currencies_prices += sub_task_currencies_prices
#             errors.update(sub_task_errors)
#
#     instrument_for_manual_price = [int(i_id) for i_id, task_id in instrument_task.items() if task_id is None]
#     _l.debug('instrument_for_manual_price: %s', instrument_for_manual_price)
#     manual_instruments_prices, manual_instruments_errors = _create_instrument_manual_prices(
#         options=options, instruments=instrument_for_manual_price)
#
#     instruments_prices += manual_instruments_prices
#     errors.update(manual_instruments_errors)
#
#     instrument_for_default_price = [int(i_id) for i_id, task_id in instrument_task.items() if
#                                     task_id == 'instrument_default']
#     _l.debug('instrument_for_default_price: %s', instrument_for_default_price)
#
#     default_instruments_prices, default_instruments_errors = _create_instrument_default_prices(
#         options=options, instruments=instrument_for_default_price, pricing_policies=pricing_policies)
#
#     instruments_prices += default_instruments_prices
#     errors.update(default_instruments_errors)
#
#     currencies_for_default_price = [int(i_id) for i_id, task_id in currency_task.items() if
#                                     task_id == 'currency_default']
#     _l.debug('currencies_for_default_price: %s', currencies_for_default_price)
#
#     default_currencies_prices, default_currencies_errors = _create_currency_default_prices(options=options,
#                                                                                            currencies=currencies_for_default_price,
#                                                                                            pricing_policies=pricing_policies)
#
#     currencies_prices += default_currencies_prices
#     errors.update(default_currencies_errors)
#
#     if errors:
#         options['errors'] = errors
#         task.options_object = options
#         task.result_object = result
#         task.status = Task.STATUS_ERROR
#         task.save()
#
#         celery_task.task_status = task.STATUS_ERROR
#         celery_task.save()
#
#     if fill_days > 0:
#         fill_date_from = date_to + timedelta(days=1)
#         instrument_last_price = [p for p in instruments_prices if p.date == date_to]
#         _l.debug('instrument last prices: %s', instrument_last_price)
#         for p in instrument_last_price:
#             instruments_prices + fill_instrument_price(fill_date_from, fill_days, p)
#
#         currency_last_price = [p for p in currencies_prices if p.date == date_to]
#         _l.debug('currency last prices: %s', currency_last_price)
#         for p in currency_last_price:
#             currencies_prices += fill_currency_price(fill_date_from, fill_days, p)
#
#     _l.debug('instruments_prices: %s', instruments_prices)
#     _l.debug('currencies_prices: %s', currencies_prices)
#
#     for p in instruments_prices:
#         # p.calculate_accrued_price(save=False)
#         accrued_price = p.instrument.get_accrued_price(p.date)
#         p.accrued_price = accrued_price if accrued_price is not None else 0.0
#
#     with transaction.atomic():
#         _l.debug('instruments_pk: %s', instruments_pk)
#         existed_instrument_prices = {
#             (p.instrument_id, p.pricing_policy_id, p.date): p
#             for p in PriceHistory.objects.filter(instrument__in=instruments_pk,
#                                                  date__range=(date_from, date_to + timedelta(days=fill_days)))
#         }
#         _l.debug('existed_instrument_prices: %s', existed_instrument_prices)
#         for p in instruments_prices:
#             op = existed_instrument_prices.get((p.instrument_id, p.pricing_policy_id, p.date), None)
#             if op is None:
#                 p.save()
#             else:
#                 if override_existed:
#                     op.principal_price = p.principal_price
#                     op.accrued_price = p.accrued_price
#                     op.save()
#
#         _l.debug('currencies_pk: %s', currencies_pk)
#         existed_currency_prices = {
#             (p.currency_id, p.pricing_policy_id, p.date): p
#             for p in CurrencyHistory.objects.filter(currency__in=currencies_pk,
#                                                     date__range=(date_from, date_to + timedelta(days=fill_days)))
#         }
#         _l.debug('existed_currency_prices: %s', existed_currency_prices)
#         for p in currencies_prices:
#             op = existed_currency_prices.get((p.currency_id, p.pricing_policy_id, p.date), None)
#
#             if op is None:
#                 p.save()
#             else:
#                 if override_existed:
#                     op.fx_rate = p.fx_rate
#                     op.save()
#
#         if is_yesterday:
#             instrument_price_real = {(p.instrument_id, p.pricing_policy_id) for p in instruments_prices
#                                      if p.date == date_to}
#             currency_price_real = {(p.currency_id, p.pricing_policy_id) for p in currencies_prices
#                                    if p.date == date_to}
#
#             instrument_price_expected = set()
#             currency_price_expected = set()
#             for pp in pricing_policies:
#                 for i_id, task_id in instrument_task.items():
#                     instrument_price_expected.add((int(i_id), pp.id))
#
#                 for c_id, task_id in currency_task.items():
#                     currency_price_expected.add((int(c_id), pp.id))
#
#             instrument_price_missed = instrument_price_expected.difference(instrument_price_real)
#             # instrument_price_missed_objects = []
#             # for instrument_id, pricing_policy_id in instrument_price_missed:
#             #     op = existed_instrument_prices.get((instrument_id, pricing_policy_id, date_to), None)
#             #     if op is None:
#             #         op = PriceHistory(instrument_id=instrument_id, pricing_policy_id=pricing_policy_id, date=date_to)
#             #     instrument_price_missed_objects.append(op)
#             # instrument_price_missed = PriceHistorySerializer(instance=instrument_price_missed_objects, many=True,
#             #                                                  context={'member': task.member}).data
#             result['instrument_price_missed'] = list(instrument_price_missed)
#
#             currency_price_missed = currency_price_expected.difference(currency_price_real)
#             # currency_price_missed_objects = []
#             # for currency_id, pricing_policy_id in currency_price_missed:
#             #     op = existed_currency_prices.get((currency_id, pricing_policy_id, date_to), None)
#             #     if op is None:
#             #         op = CurrencyHistory(currency_id=currency_id, pricing_policy_id=pricing_policy_id, date=date_to)
#             #     currency_price_missed_objects.append(op)
#             # currency_price_missed = CurrencyHistorySerializer(instance=currency_price_missed_objects, many=True).data
#             result['currency_price_missed'] = list(currency_price_missed)
#
#             _l.debug('instrument_price_missed: %s', instrument_price_missed)
#             _l.debug('currency_price_missed: %s', currency_price_missed)
#
#         task.options_object = options
#         task.result_object = result
#         task.status = Task.STATUS_DONE
#         task.save()
#
#         celery_task.task_status = Task.STATUS_DONE
#         celery_task.save()
#
#     return task_id


# DEPRECATED SINCE 22.09.2020 DELETE SOON
def _create_currency_default_prices(options, currencies, pricing_policies):
    _l.debug('create_currency_default_prices: currencies=%s', currencies)

    errors = {}
    prices = []

    date_from = parse_date_iso(options['date_from'])
    date_to = parse_date_iso(options['date_to'])

    days = (date_to - date_from).days + 1

    for c in Currency.objects.filter(pk__in=currencies):

        for pp in pricing_policies:

            for d in rrule(freq=DAILY, count=days, dtstart=date_from):
                price = CurrencyHistory(
                    currency=c,
                    pricing_policy=pp,
                    date=d.date(),
                    fx_rate=c.default_fx_rate
                )

                prices.append(price)

    return prices, errors


# DEPRECATED SINCE 22.09.2020 DELETE SOON
def _create_instrument_default_prices(options, instruments, pricing_policies):
    _l.debug('create_instrument_default_prices: instruments=%s', instruments)

    date_from = parse_date_iso(options['date_from'])
    date_to = parse_date_iso(options['date_to'])

    errors = {}
    prices = []

    days = (date_to - date_from).days + 1

    for i in Instrument.objects.filter(pk__in=instruments):

        for pp in pricing_policies:

            for dt in rrule(freq=DAILY, count=days, dtstart=date_from):
                d = dt.date()
                price = PriceHistory(
                    instrument=i,
                    pricing_policy=pp,
                    date=d,
                    principal_price=i.default_price
                )

                prices.append(price)

    return prices, errors


# DEPRECATED SINCE 22.09.2020 DELETE SOON
def _create_instrument_manual_prices(options, instruments):
    _l.debug('create_instrument_manual_prices: instruments=%s', instruments)

    date_from = parse_date_iso(options['date_from'])
    date_to = parse_date_iso(options['date_to'])
    is_yesterday = options['is_yesterday']
    fill_days = options['fill_days']

    errors = {}
    prices = []

    if is_yesterday:
        for i in Instrument.objects.filter(pk__in=instruments):
            for mf in i.manual_pricing_formulas.all():
                if mf.expr:
                    values = {
                        'd': date_to
                    }
                    try:
                        principal_price = formula.safe_eval(mf.expr, names=values)
                    except formula.InvalidExpression:
                        AbstractProvider.fail_manual_pricing_formula(errors, mf, values)
                        continue
                    price = PriceHistory(
                        instrument=i,
                        pricing_policy=mf.pricing_policy,
                        date=date_to,
                        principal_price=principal_price
                    )
                    prices.append(price)
    else:
        days = (date_to - date_from).days + 1

        for i in Instrument.objects.filter(pk__in=instruments):
            safe_instrument = {
                'id': i.id,
            }
            for mf in i.manual_pricing_formulas.all():
                if mf.expr:
                    for dt in rrule(freq=DAILY, count=days, dtstart=date_from):
                        d = dt.date()
                        values = {
                            'd': d,
                            'instrument': safe_instrument,
                        }
                        try:
                            principal_price = formula.safe_eval(mf.expr, names=values)
                        except formula.InvalidExpression:
                            AbstractProvider.fail_manual_pricing_formula(errors, mf, values)
                            continue
                        price = PriceHistory(
                            instrument=i,
                            pricing_policy=mf.pricing_policy,
                            date=d,
                            principal_price=principal_price
                        )
                        prices.append(price)

    return prices, errors


def test_certificate(master_user=None, member=None, task=None):
    _l.debug('test_certificate: master_user_id=%s, task=%s',
            getattr(master_user, 'id', None), getattr(task, 'info', None))

    try:
        if task is None:
            with transaction.atomic():

                options = {

                }

                task = Task(
                    master_user=master_user,
                    member=member,
                    provider_id=1,
                    status=Task.STATUS_PENDING,
                    action=Task.ACTION_PRICING
                )

                task.options_object = options
                task.save()

                transaction.on_commit(lambda: test_certificate_async.apply_async(kwargs={'task_id': task.id}))

            return task, False
        else:
            if task.status == Task.STATUS_DONE:
                return task, True
            return task, False

    except Exception as e:

        _l.info('test_certificate error %s ' % e)
        _l.info(traceback.print_exc())

        return task, False


# DEPRECATED SINCE 22.09.2020 DELETE SOON
# def download_pricing(master_user=None, member=None, date_from=None, date_to=None, is_yesterday=None, balance_date=None,
#                      fill_days=None, override_existed=None, task=None):
#     _l.info('download pricing')
#
#     _l.debug('download_pricing: master_user_id=%s, task=%s, date_from=%s, date_to=%s, is_yesterday=%s,'
#             ' balance_date=%s, fill_days=%s, override_existed=%s',
#             getattr(master_user, 'id', None), getattr(task, 'info', None), date_from, date_to, is_yesterday,
#             balance_date, fill_days, override_existed)
#     if task is None:
#         with transaction.atomic():
#             options = {
#                 'date_from': date_from,
#                 'date_to': date_to,
#                 'is_yesterday': is_yesterday,
#                 'balance_date': balance_date,
#                 'fill_days': fill_days,
#                 'override_existed': override_existed,
#             }
#             task = Task(
#                 master_user=master_user,
#                 member=member,
#                 provider_id=None,
#                 status=Task.STATUS_PENDING,
#                 action=Task.ACTION_PRICING
#             )
#             task.options_object = options
#             task.save()
#
#             if member:
#                 celery_task = CeleryTask.objects.create(master_user=master_user,
#                                                         member=member,
#                                                         task_status=task.status,
#                                                         started_at=datetime_now(),
#                                                         task_type='user_download_pricing', task_id=task.id)
#             else:
#                 celery_task = CeleryTask.objects.create(master_user=master_user,
#                                                         task_status=task.status,
#                                                         started_at=datetime_now(),
#                                                         task_type='automated_download_pricing', task_id=task.id)
#
#             celery_task.save()
#
#             # transaction.on_commit(lambda: download_pricing_async.apply_async(kwargs={'task_id': task.id}, countdown=1))
#             transaction.on_commit(lambda: download_pricing_async.apply_async(kwargs={'task_id': task.id}))
#         return task, False
#     else:
#         if task.status == Task.STATUS_DONE:
#
#             try:
#                 celery_task = CeleryTask.objects.get(master_user=master_user, task_id=task.id,
#                                                      task_type='user_download_pricing')
#
#                 celery_task.task_status = Task.STATUS_DONE
#                 celery_task.save()
#
#             except CeleryTask.DoesNotExist:
#                 celery_task = None
#
#             return task, True
#         return task, False


# @shared_task(name='integrations.file_import_delete', ignore_result=True)
# def file_import_delete_async(path):
#     _l.debug('file_import_delete_async: path=%s', path)
#     import_file_storage.delete(path)
#
#
# def schedule_file_import_delete(path, countdown=None):
#     if countdown == 0:
#         file_import_delete_async(path=path)
#     else:
#         if not getattr(settings, 'CELERY_TASK_ALWAYS_EAGER', False):
#             countdown = countdown or 600
#             _l.debug('schedule_file_import_delete: path=%s, countdown=%s', path, countdown)
#             file_import_delete_async.apply_async(kwargs={'path': path}, countdown=countdown)

def generate_file_report(result_object, master_user, scheme, type, name, context=None):


    def get_unique_columns(res_object):

        unique_columns = []

        for item in res_object['error_rows']:

            for item_column in item['error_data']['columns']['executed_input_expressions']:

                column = item_column + ':' + item['error_data']['data']['transaction_type_selector'][0]

                if column not in unique_columns:
                    unique_columns.append(column)

        return unique_columns

    def generate_columns_for_file(res_object):

        columns = ['Row number']

        # _l.debug('res_object %s' % res_object)

        if len(res_object['error_rows']):

            columns = columns + res_object['error_rows'][0]['error_data']['columns']['imported_columns']
            columns = columns + res_object['error_rows'][0]['error_data']['columns']['converted_imported_columns']
            columns = columns + res_object['error_rows'][0]['error_data']['columns']['calculated_columns']
            columns = columns + res_object['error_rows'][0]['error_data']['columns']['transaction_type_selector']

            unique_columns = get_unique_columns(res_object)

            for unique_column in unique_columns:
                columns.append(unique_column)

        columns.append('Error Message')
        columns.append('Reaction')

        return columns

    def generate_columns_data_for_file(instance, error_row):

        result = []
        unique_columns = get_unique_columns(instance)

        index = 0

        for unique_column in unique_columns:

            result.append('')  # result[index] = ''

            item_column_index = 0

            for item_column in error_row['error_data']['columns']['executed_input_expressions']:

                column = item_column + ':' + error_row['error_data']['data']['transaction_type_selector'][0]

                if column == unique_column:

                    if error_row['error_data']['data']['executed_input_expressions'][item_column_index]:
                        result[index] = error_row['error_data']['data']['executed_input_expressions'][item_column_index]

                item_column_index = item_column_index + 1

            index = index + 1

        return result

    _l.info('generate_file_report error_handler %s' % scheme.error_handler)
    _l.info('generate_file_report missing_data_handler %s' % scheme.missing_data_handler)

    result = []
    error_rows = []

    for item in result_object['error_rows']:
        if item['level'] == 'error':
            error_rows.append(item)

    result.append('Type, ' + name)
    result.append('Scheme, ' + scheme.user_code)
    result.append('Error handle, ' + scheme.error_handler)
    # result.append('Filename, ' + instance.file.name)
    result.append('Import Rules - if object is not found, ' + scheme.missing_data_handler)

    rowsSuccessCount = 0

    if scheme.error_handler == 'break':
        if 'error_row_index' in result_object and result_object['error_row_index']:
            rowsSuccessCount = result_object['error_row_index'] - 1
        else:
            rowsSuccessCount = result_object['total_rows'] - len(error_rows)
    else:
        rowsSuccessCount = result_object['total_rows'] - len(error_rows)

    result.append('Rows total, ' + str(result_object['total_rows']))
    result.append('Rows success import, ' + str(rowsSuccessCount))
    result.append('Rows fail import, ' + str(len(error_rows)))

    columns = generate_columns_for_file(result_object)

    column_row_list = []

    for item in columns:
        column_row_list.append('"' + str(item) + '"')

    column_row = ','.join(column_row_list)

    result.append(column_row)

    for error_row in result_object['error_rows']:

        content = []

        content.append(error_row['original_row_index'])

        content = content + error_row['error_data']['data']['imported_columns']
        content = content + error_row['error_data']['data']['converted_imported_columns']
        content = content + error_row['error_data']['data']['calculated_columns']
        content = content + error_row['error_data']['data']['transaction_type_selector']
        content = content + generate_columns_data_for_file(result_object, error_row)

        content.append(error_row['error_message'])
        content.append(error_row['error_reaction'])

        content_row_list = []

        for item in content:
            content_row_list.append('"' + str(item) + '"')

        content_row = ','.join(content_row_list)

        result.append(content_row)

    result = '\n'.join(result)

    current_date_time = now().strftime("%Y-%m-%d-%H-%M")

    file_name = 'file_report_%s.csv' % current_date_time

    file_report = FileReport()

    _l.debug('generate_file_report uploading file ')

    file_report.upload_file(file_name=file_name, text=result, master_user=master_user)
    file_report.master_user = master_user
    file_report.name = "%s %s" % (name, current_date_time)
    file_report.file_name = file_name
    file_report.type = type
    file_report.notes = 'System File'

    file_report.save()

    _l.debug('file_report %s' % file_report)
    _l.debug('file_report %s' % file_report.file_url)

    return file_report.pk


def generate_file_report_old(instance, master_user, type, name, context=None):
    def get_unique_columns(instance):

        unique_columns = []

        for item in instance.error_rows:

            for item_column in item['error_data']['columns']['executed_input_expressions']:

                column = item_column + ':' + item['error_data']['data']['transaction_type_selector'][0];

                if column not in unique_columns:
                    unique_columns.append(column)

        return unique_columns

    def generate_columns_for_file(instance):

        columns = ['Row number']

        _l.debug('instance %s' % instance)

        if len(instance.error_rows):

            columns = columns + instance.error_rows[0]['error_data']['columns']['imported_columns']
            columns = columns + instance.error_rows[0]['error_data']['columns']['converted_imported_columns']
            columns = columns + instance.error_rows[0]['error_data']['columns']['calculated_columns']
            columns = columns + instance.error_rows[0]['error_data']['columns']['transaction_type_selector']

            unique_columns = get_unique_columns(instance)

            for unique_column in unique_columns:
                columns.append(unique_column)

        columns.append('Error Message')
        columns.append('Reaction')

        return columns

    def generate_columns_data_for_file(instance, error_row):

        result = []
        unique_columns = get_unique_columns(instance)

        index = 0

        for unique_column in unique_columns:

            result.append('')  # result[index] = ''

            item_column_index = 0

            for item_column in error_row['error_data']['columns']['executed_input_expressions']:

                column = item_column + ':' + error_row['error_data']['data']['transaction_type_selector'][0]

                if column == unique_column:

                    if error_row['error_data']['data']['executed_input_expressions'][item_column_index]:
                        result[index] = error_row['error_data']['data']['executed_input_expressions'][item_column_index]

                item_column_index = item_column_index + 1

            index = index + 1

        return result

    result = []
    error_rows = []

    for item in instance.error_rows:
        if item['level'] == 'error':
            error_rows.append(item)

    result.append('Type, ' + 'Transaction Import')
    result.append('Error handle, ' + instance.error_handling)
    # result.append('Filename, ' + instance.file.name)
    result.append('Import Rules - if object is not found, ' + instance.missing_data_handler);

    rowsSuccessCount = 0

    if instance.error_handling == 'break':
        if instance.error_row_index:
            rowsSuccessCount = instance.error_row_index - 1
        else:
            rowsSuccessCount = instance.total_rows - len(error_rows)
    else:
        rowsSuccessCount = instance.total_rows - len(error_rows)

    result.append('Rows total, ' + str(instance.total_rows))
    result.append('Rows success import, ' + str(rowsSuccessCount))
    result.append('Rows fail import, ' + str(len(error_rows)))

    columns = generate_columns_for_file(instance)

    column_row_list = []

    for item in columns:
        column_row_list.append('"' + str(item) + '"')

    column_row = ','.join(column_row_list)

    result.append(column_row)

    for error_row in instance.error_rows:

        content = []

        content.append(error_row['original_row_index'])

        content = content + error_row['error_data']['data']['imported_columns']
        content = content + error_row['error_data']['data']['converted_imported_columns']
        content = content + error_row['error_data']['data']['calculated_columns']
        content = content + error_row['error_data']['data']['transaction_type_selector']
        content = content + generate_columns_data_for_file(instance, error_row)

        content.append(error_row['error_message'])
        content.append(error_row['error_reaction'])

        content_row_list = []

        for item in content:
            content_row_list.append('"' + str(item) + '"')

        content_row = ','.join(content_row_list)

        result.append(content_row)

    result = '\n'.join(result)

    current_date_time = now().strftime("%Y-%m-%d-%H-%M")

    file_name = 'file_report_%s.csv' % current_date_time

    file_report = FileReport()

    _l.debug('generate_file_report uploading file ')

    file_report.upload_file(file_name=file_name, text=result, master_user=master_user)
    file_report.master_user = master_user
    file_report.name = "%s %s" % (name, current_date_time)
    file_report.file_name = file_name
    file_report.type = type
    file_report.notes = 'System File'

    file_report.save()

    _l.debug('file_report %s' % file_report)
    _l.debug('file_report %s' % file_report.file_url)

    return file_report.pk


@shared_task(name="integrations.complex_transaction_csv_file_import_parallel_finish", bind=True)
def complex_transaction_csv_file_import_parallel_finish(self, task_id):
    try:
        _l.info('complex_transaction_csv_file_import_parallel_finish task_id %s ' % task_id)

        celery_task = CeleryTask.objects.get(pk=task_id)

        scheme = ComplexTransactionImportScheme.objects.get(pk=celery_task.options_object['scheme_id'])

        master_user = celery_task.master_user
        member = celery_task.member

        result_object = {
            'error_rows': [],
            'total_rows': celery_task.options_object['total_rows'],
            'processed_rows': 0

        }

        _l.info('complex_transaction_csv_file_import_parallel_finish iterating over %s childs' % len(celery_task.children.all()))

        for sub_task in celery_task.children.all():

            if sub_task.result_object:

                if 'error_rows' in sub_task.result_object:

                    result_object['error_rows'] = result_object['error_rows'] + sub_task.result_object['error_rows']

                if 'processed_rows' in sub_task.result_object:

                    result_object['processed_rows'] = result_object['processed_rows'] + sub_task.result_object['processed_rows']

        result_object['stats_file_report'] = generate_file_report(result_object, master_user, scheme, 'transaction_import.import',
                                                          'Transaction Import', celery_task.options_object['execution_context'])

        if celery_task.options_object['execution_context'] and celery_task.options_object['execution_context']["started_by"] == 'procedure':

            _l.info('complex_transaction_csv_file_import_parallel_finish send final import message')

            send_system_message(master_user=celery_task.master_user,
                                source="Transaction Import Service",
                                text="Import Finished",
                                file_report_id=result_object['stats_file_report'])
        else:

            send_system_message(master_user=celery_task.master_user,
                                source="Transaction Import Service",
                                text="User %s Transaction Import Finished" % celery_task.member.username,
                                file_report_id=result_object['stats_file_report'])

        # TODO Generate File Report Here

        send_websocket_message(data={
            'type': 'transaction_import_status',
            'payload': {'task_id': task_id,
                        'state': Task.STATUS_DONE,
                        'error_rows': result_object['error_rows'],
                        'total_rows': result_object['total_rows'],
                        'processed_rows': result_object['processed_rows'],
                        'stats_file_report': result_object['stats_file_report'],
                        'scheme': scheme.id,
                        'scheme_object': {
                            'id': scheme.id,
                            'scheme_name': scheme.user_code,
                            'delimiter': scheme.delimiter,
                            'error_handler': scheme.error_handler,
                            'missing_data_handler': scheme.missing_data_handler
                        }}
        }, level="member",
            context={"master_user": master_user, "member": member})

        celery_task.result_object = result_object

        celery_task.status = CeleryTask.STATUS_DONE
        celery_task.save()

    except Exception as e:

        _l.info('Exception occurred %s' % e)
        _l.info(traceback.format_exc())


@shared_task(name='integrations.complex_transaction_csv_file_import', bind=True)
def complex_transaction_csv_file_import(self, task_id):

    try:

        from poms.transactions.models import TransactionTypeInput
        from poms.integrations.serializers import ComplexTransactionCsvFileImport

        celery_task = CeleryTask.objects.get(pk=task_id)
        parent_celery_task = celery_task.parent

        celery_task.status = CeleryTask.STATUS_PENDING
        celery_task.save()

        master_user = celery_task.master_user
        member = celery_task.member

        instance = ComplexTransactionCsvFileImport(task_id=task_id, master_user=master_user, member=member, skip_first_line=True)

        scheme = ComplexTransactionImportScheme.objects.get(pk=celery_task.options_object['scheme_id'])

        instance.scheme = scheme
        instance.error_handling = scheme.error_handler
        instance.delimiter = scheme.delimiter
        instance.missing_data_handler = scheme.missing_data_handler
        instance.file_path = celery_task.options_object['file_path']
        execution_context = celery_task.options_object['execution_context']

        instance.processed_rows = 0

        scheme = instance.scheme
        scheme_inputs = list(scheme.inputs.all())
        scheme_calculated_inputs = list(scheme.calculated_inputs.all())
        scheme_selector_values = list(scheme.selector_values.all())

        master_user = instance.master_user
        member = instance.member

        rule_scenarios = scheme.rule_scenarios.prefetch_related('transaction_type', 'fields',
                                                                'fields__transaction_type_input').all()

        _l.info('scheme %s - inputs=%s, rules=%s', scheme,
                            [(i.name, i.column) for i in scheme_inputs],
                            [(r.transaction_type.user_code) for r in rule_scenarios])


        _l.info('scheme %s - column_matcher %s', (scheme, scheme.column_matcher))
        _l.info('scheme %s - has_header_row %s', (scheme, scheme.has_header_row))

        default_rule_scenario = None

        for scenario in rule_scenarios:
            if scenario.is_default_rule_scenario:
                default_rule_scenario = scenario

        mapping_map = {
            Account: AccountMapping,
            Currency: CurrencyMapping,
            Instrument: InstrumentMapping,
            InstrumentType: InstrumentTypeMapping,
            Counterparty: CounterpartyMapping,
            Responsible: ResponsibleMapping,
            Strategy1: Strategy1Mapping,
            Strategy2: Strategy2Mapping,
            Strategy3: Strategy3Mapping,
            DailyPricingModel: DailyPricingModelMapping,
            PaymentSizeDetail: PaymentSizeDetailMapping,
            Portfolio: PortfolioMapping,
            PriceDownloadScheme: PriceDownloadSchemeMapping,
            Periodicity: PeriodicityMapping,
            AccrualCalculationModel: AccrualCalculationModelMapping,

        }

        props_map = {
            Account: 'account',
            Currency: 'currency',
            Instrument: 'instrument',
            InstrumentType: 'instrument_type',
            Counterparty: 'counterparty',
            Responsible: 'responsible',
            Strategy1: 'strategy1',
            Strategy2: 'strategy2',
            Strategy3: 'strategy3',
            DailyPricingModel: 'daily_pricing_model',
            PaymentSizeDetail: 'payment_size_detail',
            Portfolio: 'portfolio',
            PriceDownloadScheme: 'price_download_scheme',
            Periodicity: 'periodicity',
            AccrualCalculationModel: 'accrual_calculation_model',
        }

        def _get_default_relation(field):

            i = field.transaction_type_input

            model_class = i.content_type.model_class()
            model_map_class = mapping_map[model_class]

            key = props_map[model_class]

            v = None

            ecosystem_default = EcosystemDefault.objects.get(master_user=instance.master_user)

            # _l.info('key %s' % key)
            # _l.info('value %s' % value)

            if hasattr(ecosystem_default, key):
                v = getattr(ecosystem_default, key)
            else:
                v = model_map_class.objects.get(master_user=instance.master_user, value='-').content_object

            return v

        def _convert_value(field, value, error_rows):
            i = field.transaction_type_input

            if i.value_type == TransactionTypeInput.STRING:
                return str(value)

            if i.value_type == TransactionTypeInput.SELECTOR:
                return str(value)

            elif i.value_type == TransactionTypeInput.NUMBER:
                return float(value)

            elif i.value_type == TransactionTypeInput.DATE:
                if not isinstance(value, date):
                    return formula._parse_date(value)
                else:
                    return value

            elif i.value_type == TransactionTypeInput.RELATION:
                model_class = i.content_type.model_class()
                model_map_class = mapping_map[model_class]

                v = None

                try:
                    v = model_map_class.objects.get(master_user=instance.master_user, value=value).content_object
                except Exception:

                    try:

                        v = model_class.objects.get(master_user=instance.master_user, user_code=value)

                    except (model_class.DoesNotExist, KeyError):
                        v = None

                        _l.info("User code %s not found for %s " % (value, field.transaction_type_input.name))

                if not v:

                    if instance.missing_data_handler == 'set_defaults':

                        v = _get_default_relation(field)

                    else:
                        error_rows['error_message'] = error_rows[
                                                          'error_message'] + ' Can\'t find relation of ' + \
                                                      '[' + field.transaction_type_input.name + ']' + '(value:' + \
                                                      value + ')'

                return v

        def update_row_with_calculated_data(row, inputs):

            for i in scheme_calculated_inputs:

                try:
                    value = formula.safe_eval(i.name_expr, names=inputs)
                    row.append(value)

                except Exception:
                    _l.debug('can\'t process calculated input: %s|%s', i.name, i.column, exc_info=True)
                    row.append("Invalid Expression")

            return row

        def _process_rule_scenario(processed_scenarios, scheme_rule, inputs, error_rows, row_index):

            _l.info('_process_rule_scenario %s %s ' % (row_index, scheme_rule))

            result = None

            processed_scenarios = processed_scenarios + 1

            rule = scheme_rule

            fields = {}
            fields_error = []
            for field in rule.fields.all():
                try:
                    field_value = formula.safe_eval(field.value_expr, names=inputs)
                    field_value = _convert_value(field, field_value, error_rows)
                    fields[field.transaction_type_input.name] = field_value

                except (Exception, ValueError, ExpressionEvalError):
                    _l.debug('can\'t process field: %s|%s', field.transaction_type_input.name,
                             field.transaction_type_input.pk)
                    fields_error.append(field)


            _l.debug('fields (step 1): error=%s, values=%s', fields_error, fields)

            if fields_error:

                error_rows['level'] = 'error'

                error_rows['error_message'] = error_rows['error_message'] + '\n' + str(ugettext(
                    'Can\'t process fields: %(fields)s') % {
                                                                                           'fields': ', '.join(
                                                                                               '[' + f.transaction_type_input.name + '] '
                                                                                               for f in
                                                                                               fields_error)
                                                                                       })

                if instance.break_on_error:
                    instance.error_row_index = row_index
                    error_rows['error_reaction'] = 'Break'
                    instance.error_rows.append(error_rows)

                    result = 'break'

                    return result, processed_scenarios
                else:
                    error_rows['error_reaction'] = 'Continue import'

                    result = 'continue'

                    return result, processed_scenarios

            with transaction.atomic():
                try:
                    tt_process = TransactionTypeProcess(
                        transaction_type=rule.transaction_type,
                        default_values=fields,
                        context={
                            'master_user': instance.master_user,
                            'member': instance.member,
                        },
                        uniqueness_reaction=instance.scheme.book_uniqueness_settings,
                        member=instance.member
                    )
                    tt_process.process()

                    _l.debug('tt_process %s' % tt_process)

                    if tt_process.uniqueness_status == 'skip':
                        error_rows['level'] = 'skip'
                        error_rows['error_message'] = error_rows['error_message'] + str(
                            ugettext('Unique code already exist. Skip'))

                    if tt_process.uniqueness_status == 'error':
                        error_rows['level'] = 'error'
                        error_rows['error_message'] = error_rows['error_message'] + str(
                            ugettext('Unique code already exist. Error'))

                    processed_scenarios = processed_scenarios + 1

                except Exception as e:

                    error_rows['level'] = 'error'

                    _l.debug("can't process transaction type", exc_info=True)

                    _l.debug('error %s' % e)

                    transaction.set_rollback(True)
                    if instance.break_on_error:
                        instance.error_row_index = row_index
                        error_rows['error_reaction'] = 'Break'
                        result = 'break'
                        return result, processed_scenarios
                    else:
                        result = 'continue'
                        return result, processed_scenarios
                finally:
                    _l.debug("final")
                    # if settings.DEBUG:
                    #     transaction.set_rollback(True)

            return result, processed_scenarios

        def _process_csv_file(file, original_file, original_file_path):

            _l.info('_process_csv_file %s ' % instance.file_path)

            instance.processed_rows = 0

            reader = []

            if '.csv' in instance.file_path or (execution_context and execution_context["started_by"] == 'procedure'):

                delimiter = instance.delimiter.encode('utf-8').decode('unicode_escape')

                reader = csv.reader(file, delimiter=delimiter, quotechar=instance.quotechar,
                                    strict=False, skipinitialspace=True)

            elif '.xlsx' in instance.file_path:
                _l.info('trying to parse excel %s ' % instance.file_path)

                wb = load_workbook(filename=original_file_path)

                if instance.scheme.spreadsheet_active_tab_name and instance.scheme.spreadsheet_active_tab_name in wb.sheetnames:
                    ws = wb[instance.scheme.spreadsheet_active_tab_name]
                else:
                    ws = wb.active

                _l.info('ws %s' % ws)
                _l.info('task_instance.scheme.spreadsheet_start_cell %s' % instance.scheme.spreadsheet_start_cell)

                reader = []

                if instance.scheme.spreadsheet_start_cell == 'A1':

                    for r in ws.rows:
                        reader.append([cell.value for cell in r])

                else:

                    start_cell_row_number = int(re.search(r'\d+', instance.scheme.spreadsheet_start_cell)[0])
                    start_cell_letter = instance.scheme.spreadsheet_start_cell.split(str(start_cell_row_number))[0]

                    start_cell_column_number = column_index_from_string(start_cell_letter)


                    row_number = 1

                    for r in ws.rows:

                        row_values = []

                        if row_number >= start_cell_row_number:

                            for cell in r:


                                if cell.column >= start_cell_column_number:
                                    row_values.append(cell.value)

                            reader.append(row_values)

                        row_number = row_number + 1

            first_row = None



            # reader = [{}]

            _process_list_of_items(reader)

        def _process_list_of_items(items):

            input_column_name_map = {}

            for row_index, row in enumerate(items):

                _l.info('process row_index %s ' % row_index)
                _l.info('process row %s ' % row)

                if row_index == 0:
                    first_row = row

                    _local_index = 0
                    for item in first_row:

                        input_column_name_map[item] = _local_index
                        _local_index = _local_index + 1


                # _l.debug('process row: %s -> %s', row_index, row)
                if (row_index == 0 and instance.skip_first_line and not scheme.has_header_row) or not row:
                    _l.debug('skip first row')
                    continue

                inputs_raw = {}
                inputs = {}
                inputs_error = []
                inputs_conversion_error = []
                calculated_columns_error = []

                error_rows = {
                    'level': 'info',
                    'error_message': '',
                    'inputs': inputs_raw,
                    'original_row_index': row_index,
                    'original_row': row,
                    'error_data': {
                        'columns': {
                            'imported_columns': [],
                            'calculated_columns': [],
                            'converted_imported_columns': [],
                            'transaction_type_selector': [],
                            'executed_input_expressions': []
                        },
                        'data': {
                            'imported_columns': [],
                            'calculated_columns': [],
                            'converted_imported_columns': [],
                            'transaction_type_selector': [],
                            'executed_input_expressions': []
                        }

                    },
                    'error_reaction': "Success"
                }



                for i in scheme_inputs:

                    error_rows['error_data']['columns']['imported_columns'].append(i.name)


                    if instance.scheme.column_matcher == 'index':
                        try:
                            inputs_raw[i.name] = row[i.column - 1]
                            error_rows['error_data']['data']['imported_columns'].append(row[i.column - 1])
                        except Exception:
                            _l.debug('can\'t process input: %s|%s', i.name, i.column, exc_info=True)
                            _l.debug('can\'t process inputs_raw: %s|%s', inputs_raw)
                            error_rows['error_data']['data']['imported_columns'].append(ugettext('Invalid expression'))
                            inputs_error.append(i)

                    if instance.scheme.column_matcher == 'name':

                        try:
                            if type(row) is dict:
                                inputs_raw[i.name] = row[i.name]
                            else:
                                _col_index = input_column_name_map[i.name]
                                inputs_raw[i.name] = row[_col_index]
                                error_rows['error_data']['data']['imported_columns'].append(row[_col_index])
                        except Exception:
                            _l.debug('can\'t process input: %s|%s', i.name, i.column, exc_info=True)
                            _l.debug('can\'t process inputs_raw: %s|%s', inputs_raw)
                            error_rows['error_data']['data']['imported_columns'].append(ugettext('Invalid expression'))
                            inputs_error.append(i)

                # _l.debug('Row %s inputs_raw: %s' % (row_index, inputs_raw))

                if scheme.filter_expression:

                    # expr = Expression.parseString("a == 1 and b == 2")
                    expr = Expression.parseString(scheme.filter_expression)

                    if expr(inputs_raw):
                        # filter passed
                        pass
                    else:
                        _l.info("Row skipped due filter %s" % row_index)
                        continue


                original_columns_count = len(row)

                if inputs_error:

                    error_rows['level'] = 'error'

                    error_rows['error_message'] = error_rows['error_message'] + str(
                        ugettext('Can\'t process fields: %(inputs)s') % {
                            'inputs': ', '.join('[' + i.name + '] (Can\'t find input)' for i in inputs_error)
                        })
                    instance.error_rows.append(error_rows)
                    if instance.break_on_error:
                        error_rows['error_reaction'] = 'Break'
                        instance.error_row_index = row_index
                        instance.error_rows.append(error_rows)
                        return
                    else:
                        error_rows['error_reaction'] = 'Continue import'
                        continue

                for i in scheme_inputs:

                    error_rows['error_data']['columns']['converted_imported_columns'].append(
                        i.name + ': Conversion Expression ' + '(' + i.name_expr + ')')

                    try:
                        inputs[i.name] = formula.safe_eval(i.name_expr, names=inputs_raw)
                        error_rows['error_data']['data']['converted_imported_columns'].append(row[i.column - 1])
                    except Exception:
                        _l.debug('can\'t process conversion input: %s|%s', i.name, i.column, exc_info=True)
                        error_rows['error_data']['data']['converted_imported_columns'].append(
                            ugettext('Invalid expression'))
                        inputs_conversion_error.append(i)

                # _l.debug('Row %s inputs_conversion: %s' % (row_index, inputs))

                if inputs_conversion_error:

                    error_rows['level'] = 'error'

                    error_rows['error_message'] = error_rows['error_message'] + str(
                        ugettext('Can\'t process fields: %(inputs)s') % {
                            'inputs': ', '.join(
                                '[' + i.name + '] (Imported column conversion expression, value; "' + i.name_exp + '")' for
                                i in inputs_conversion_error)
                        })
                    instance.error_rows.append(error_rows)
                    if instance.break_on_error:
                        error_rows['error_reaction'] = 'Break'
                        instance.error_row_index = row_index
                        instance.error_rows.append(error_rows)
                        return
                    else:
                        error_rows['error_reaction'] = 'Continue import'
                        continue

                row = update_row_with_calculated_data(row, inputs)

                for i in scheme_calculated_inputs:

                    error_rows['error_data']['columns']['calculated_columns'].append(i.name)

                    try:

                        index = original_columns_count + i.column - 1

                        # print('index %s ' % index)
                        # print('i.name %s ' % i.name)

                        inputs[i.name] = row[index]

                        error_rows['error_data']['data']['calculated_columns'].append(row[index])
                    except Exception:
                        _l.debug('can\'t process input: %s|%s', i.name, i.column, exc_info=True)
                        error_rows['error_data']['data']['calculated_columns'].append(ugettext('Invalid expression'))
                        calculated_columns_error.append(i)

                # _l.debug('Row %s inputs_with_calculated: %s' % (row_index, inputs))

                try:
                    rule_value = formula.safe_eval(scheme.rule_expr, names=inputs)
                except Exception:

                    error_rows['level'] = 'error'

                    _l.debug('can\'t process rule expression', exc_info=True)
                    error_rows['error_message'] = error_rows['error_message'] + '\n' + str(ugettext(
                        'Can\'t eval rule expression'))
                    instance.error_rows.append(error_rows)
                    if instance.break_on_error:
                        instance.error_row_index = row_index
                        instance.error_rows.append(error_rows)
                        return
                    else:
                        continue

                if not rule_value:

                    _l.debug('no rule value: %s', rule_value)

                    error_rows['level'] = 'error'

                    error_rows['error_message'] = error_rows['error_message'] + str(ugettext('Rule expression is invalid'))

                    if instance.break_on_error:
                        instance.error_row_index = row_index
                        error_rows['error_reaction'] = 'Break'
                        instance.error_rows.append(error_rows)
                        return
                    else:
                        error_rows['error_reaction'] = 'Continue import'
                        continue

                # else:
                #     _l.debug('rule value: %s', rule_value)

                matched_selector = False
                processed_scenarios = 0

                unknown_rule = True

                for selector in scheme_selector_values:
                    if selector.value == rule_value:
                        unknown_rule = False

                if unknown_rule and default_rule_scenario:
                    _l.info("Process rule %s with default rule scenario "  % (rule_value))
                    res, processed_scenarios = _process_rule_scenario(processed_scenarios, default_rule_scenario, inputs, error_rows, row_index)

                    # TODO refactor soon
                    if res == 'break':
                        return
                    elif res == 'continue':
                        continue

                else:

                    for scheme_rule in rule_scenarios:

                        matched_selector = False

                        selector_values = scheme_rule.selector_values.all()

                        for selector_value in selector_values:

                            if selector_value.value == rule_value:
                                matched_selector = True

                        if matched_selector:

                            res, processed_scenarios = _process_rule_scenario(processed_scenarios, scheme_rule, inputs, error_rows, row_index)

                            # TODO refactor soon
                            if res == 'break':
                                return
                            elif res == 'continue':
                                continue

                if processed_scenarios == 0:
                    error_rows['level'] = 'error'

                    error_rows['error_message'] = error_rows['error_message'] + str('Selector does not match')

                    if instance.break_on_error:
                        instance.error_row_index = row_index
                        error_rows['error_reaction'] = 'Break'
                        instance.error_rows.append(error_rows)
                        return
                    else:
                        error_rows['error_reaction'] = 'Continue import'

                instance.error_rows.append(error_rows)

                instance.processed_rows = instance.processed_rows + 1

                total_rows = 0

                if parent_celery_task:
                    total_rows = parent_celery_task.options_object['total_rows'],
                else:
                    total_rows = instance.total_rows



                send_websocket_message(data={
                    'type': 'transaction_import_status',
                    'payload': {
                        'parent_task_id': celery_task.parent_id,
                        'task_id': instance.task_id,
                        'state': Task.STATUS_PENDING,
                        'processed_rows': instance.processed_rows,
                        'parent_total_rows': total_rows,
                        'total_rows': instance.total_rows,
                        'scheme_name': scheme.user_code,
                        'file_name': instance.filename}
                }, level="member",
                    context={"master_user": master_user, "member": member})

        def _row_count_csv(file):

            delimiter = instance.delimiter.encode('utf-8').decode('unicode_escape')

            reader = csv.reader(file, delimiter=delimiter, quotechar=instance.quotechar,
                                strict=False, skipinitialspace=True)

            row_index = 0

            for row_index, row in enumerate(reader):
                pass

            _l.info("Total rows in file: %s" % row_index)

            return row_index

        def _row_count_xlsx(filename):

            wb = load_workbook(filename=filename)

            if instance.scheme.spreadsheet_active_tab_name and instance.scheme.spreadsheet_active_tab_name in wb.sheetnames:
                ws = wb[instance.scheme.spreadsheet_active_tab_name]
            else:
                ws = wb.active

            reader = []

            row_index = 0

            for r in ws.rows:
                row_index = row_index + 1

            return row_index

        instance.error_rows = []

        try:

            if celery_task.options_object and 'items' in celery_task.options_object:

                _l.info("Parse json data")

                items = celery_task.options_object['items']

                instance.total_rows = len(items)

                _process_list_of_items(items)

            else:

                _l.info("Open file %s" % instance.file_path)
                # with import_file_storage.open(instance.file_path, 'rb') as f:
                with SFS.open(instance.file_path, 'rb') as f:


                    with NamedTemporaryFile() as tmpf:

                        for chunk in f.chunks():
                            tmpf.write(chunk)
                        tmpf.flush()

                        os.link(tmpf.name, tmpf.name + '.xlsx')


                        if '.csv' in instance.file_path or (execution_context and execution_context["started_by"] == 'procedure'):

                            with open(tmpf.name, mode='rt', encoding=instance.encoding, errors='ignore') as cfr:
                                instance.total_rows = _row_count_csv(cfr)

                            with open(tmpf.name, mode='rt', encoding=instance.encoding, errors='ignore') as cf:
                                _process_csv_file(cf, f, '')

                        elif '.xlsx' in instance.file_path:

                            instance.total_rows = _row_count_xlsx(tmpf.name + '.xlsx')

                            with open(tmpf.name, mode='rt', encoding=instance.encoding, errors='ignore') as cf:
                                _process_csv_file(cf, f, tmpf.name + '.xlsx')


        except Exception:

            _l.debug('Can\'t process file', exc_info=True)
            instance.error_message = ugettext("Invalid file format or file already deleted.")

            if execution_context and execution_context["started_by"] == 'procedure':

                send_system_message(master_user=instance.master_user,
                                    source="Transaction Import Service",
                                    text="Can't process file. Possibly wrong format")

        finally:
            # import_file_storage.delete(instance.file_path)


            if celery_task.options_object and 'items' in celery_task.options_object:
                pass
            else:
                SFS.delete(instance.file_path)

            instance.error = bool(instance.error_message) or (instance.error_row_index is not None) or bool(instance.error_rows)

            # instance.stats_file_report = generate_file_report(instance, master_user, 'transaction_import.import',
            #                                                   'Transaction Import', execution_context)

            _l.debug('complex_transaction_file_import execution_context: %s', execution_context)

            # _l.debug("Reached end instance.stats_file_report: %s " % instance.stats_file_report)

            # if execution_context and execution_context["started_by"] == 'procedure':
            #
            #     _l.debug('send final import message')
            #
            #     send_system_message(master_user=instance.master_user,
            #                         source="Transaction Import Service",
            #                         text="Import Finished",
            #                         file_report_id=instance.stats_file_report)

            total_rows = 0

            if parent_celery_task:
                total_rows = parent_celery_task.options_object['total_rows'],
            else:
                total_rows = instance.total_rows

            send_websocket_message(data={
                'type': 'transaction_import_status',
                'payload': {
                            'parent_task_id': celery_task.parent_id,
                            'task_id': instance.task_id,
                            'state': Task.STATUS_DONE,
                            'processed_rows': instance.processed_rows,
                            'parent_total_rows': total_rows,
                            'total_rows': instance.total_rows,
                            'file_name': instance.filename,
                            'error_rows': instance.error_rows,
                            'stats_file_report': instance.stats_file_report,
                            'scheme': scheme.id,
                            'scheme_object': {
                                'id': scheme.id,
                                'scheme_name': scheme.user_code,
                                'delimiter': scheme.delimiter,
                                'error_handler': scheme.error_handler,
                                'missing_data_handler': scheme.missing_data_handler
                            }}
            }, level="member",
                context={"master_user": master_user, "member": member})

            result_object = {
                'processed_rows': instance.processed_rows,
                'total_rows': instance.total_rows,
                'error_row_index': instance.error_row_index,
                'file_name': instance.filename,
                'error_rows': instance.error_rows,
                'stats_file_report': instance.stats_file_report
            }

            celery_task.result_object = result_object
            celery_task.status = CeleryTask.STATUS_DONE
            celery_task.save()

        return instance
    except Exception as e:

        _l.info('Exception occurred %s' % e)
        _l.info(traceback.format_exc())


# @shared_task(name='integrations.complex_transaction_csv_file_import_parallel', bind=True)
def complex_transaction_csv_file_import_parallel(task_id):

    try:

        _l.info('complex_transaction_csv_file_import_parallel: task_id %s' % task_id)

        celery_task = CeleryTask.objects.get(pk=task_id)

        celery_task.status = CeleryTask.STATUS_PENDING
        celery_task.save()

        options_object = celery_task.options_object
        sub_tasks = []
        celery_sub_tasks = []

        if 'items' in options_object:

            sub_task = CeleryTask.objects.create(master_user=celery_task.master_user, member=celery_task.member, parent=celery_task)

            sub_task_options_object = copy.deepcopy(celery_task.options_object)

            sub_task.options_object = sub_task_options_object
            sub_task.save()

            sub_tasks.append(sub_task)

            for sub_task in sub_tasks:

                ct = complex_transaction_csv_file_import.s(task_id=sub_task.id)
                celery_sub_tasks.append(ct)


            transaction.on_commit(
                lambda:  chord(celery_sub_tasks, complex_transaction_csv_file_import_parallel_finish.si(task_id=task_id)).apply_async())


        else:

            sub_tasks = []

            lines_per_file = 300
            header_line = None

            def _get_path(master_user, file_name, ext):
                return '%s/transaction_import_files/%s.%s' % (master_user.token, file_name, ext)

            chunk = None


            with SFS.open(celery_task.options_object['file_path'], 'rb') as f:

                _l.info("Start reading file to split it into chunks")

                ext = celery_task.options_object['file_path'].split('.')[-1]

                for lineno, line in enumerate(f):

                    # _l.info('line %s' % lineno)

                    if lineno == 0:
                        header_line = line
                        # _l.info('set header line %s' % lineno)

                    if lineno % lines_per_file == 0:

                        if chunk is not None:
                            # _l.info("Saving chunk %s" % chunk)
                            SFS.save(chunk_path, chunk) # save working chunk before creating new one

                        chunk_filename = '%s_chunk_file_%s' % (celery_task.id, str(lineno) + '_' + str(lineno + lines_per_file))
                        chunk_path = _get_path(celery_task.master_user, chunk_filename, ext)

                        # _l.info('creating chunk file %s' % chunk_path)

                        chunk = BytesIO()
                        if lineno != 0:
                            chunk.write(header_line)

                        _l.info("creating chunk %s" % chunk_filename)

                        # _l.info('creating sub task for %s' % chunk_filename)

                        sub_task = CeleryTask.objects.create(master_user=celery_task.master_user, member=celery_task.member, parent=celery_task)

                        sub_task_options_object = copy.deepcopy(celery_task.options_object)
                        sub_task_options_object['file_path'] = chunk_path

                        sub_task.options_object = sub_task_options_object
                        sub_task.save()

                        sub_tasks.append(sub_task)

                    chunk.write(line)

                _l.info("Saving last chunk")
                SFS.save(chunk_path, chunk) # save working chunk before creating new one

            _l.info('sub_tasks created %s' % len(sub_tasks))
            _l.info('original file total rows %s' % lineno)

            options_object['total_rows'] = lineno

            celery_task.options_object = options_object
            celery_task.save()

            celery_sub_tasks = []

            # for sub_task in sub_tasks:
            #
            #     _l.info('initializing sub_task %s' % sub_task.options_object['file_path'])
            #
            #     ct = complex_transaction_csv_file_import.s(task_id=sub_task.id)
            #     celery_sub_tasks.append(ct)
            #
            # # chord(celery_sub_tasks, complex_transaction_csv_file_import_parallel_finish.si(task_id=task_id)).apply_async()
            # chord(celery_sub_tasks)(complex_transaction_csv_file_import_parallel_finish.si(task_id=task_id))

            for sub_task in sub_tasks:

                _l.info('initializing sub_task %s' % sub_task.options_object['file_path'])

                ct = complex_transaction_csv_file_import.s(task_id=sub_task.id)
                celery_sub_tasks.append(ct)

            _l.info('celery_sub_tasks len %s' % len(celery_sub_tasks))
            _l.info('celery_sub_tasks %s' % celery_sub_tasks)

            # chord(celery_sub_tasks, complex_transaction_csv_file_import_validate_parallel_finish.si(task_id=task_id)).apply_async()
            # chord(celery_sub_tasks)(complex_transaction_csv_file_import_parallel_finish.si(task_id=task_id))

            transaction.on_commit(
                lambda:  chord(celery_sub_tasks, complex_transaction_csv_file_import_parallel_finish.si(task_id=task_id)).apply_async())

    except Exception as e:

        _l.info('Exception occurred %s' % e)
        _l.info(traceback.format_exc())


@shared_task(name="integrations.complex_transaction_csv_file_import_validate_parallel_finish", bind=True)
def complex_transaction_csv_file_import_validate_parallel_finish(self, task_id):

    try:

        _l.info('complex_transaction_csv_file_import_validate_parallel_finish task_id %s ' % task_id)

        celery_task = CeleryTask.objects.get(pk=task_id)

        scheme = ComplexTransactionImportScheme.objects.get(pk=celery_task.options_object['scheme_id'])

        master_user = celery_task.master_user
        member = celery_task.member

        result_object = {
            'error_rows': [],
            'total_rows': celery_task.options_object['total_rows'],
            'processed_rows': 0

        }

        _l.info('complex_transaction_csv_file_import_validate_parallel_finish iterating over %s childs' % len(celery_task.children.all()))

        for sub_task in celery_task.children.all():

            if sub_task.result_object:

                if 'error_rows' in sub_task.result_object:

                    result_object['error_rows'] = result_object['error_rows'] + sub_task.result_object['error_rows']

                if 'processed_rows' in sub_task.result_object:

                    result_object['processed_rows'] = result_object['processed_rows'] + sub_task.result_object['processed_rows']

        result_object['stats_file_report'] = generate_file_report(result_object, master_user, scheme, 'transaction_import.validate',
                                                            'Transaction Import Validation', False)

        send_websocket_message(data={
            'type': 'transaction_import_status',
            'payload': {'task_id': task_id,
                        'state': Task.STATUS_DONE,
                        'error_rows': result_object['error_rows'],
                        'total_rows': result_object['total_rows'],
                        'processed_rows': result_object['processed_rows'],
                        'stats_file_report': result_object['stats_file_report'],
                        'scheme': scheme.id,
                        'scheme_object': {
                            'id': scheme.id,
                            'scheme_name': scheme.user_code,
                            'delimiter': scheme.delimiter,
                            'error_handler': scheme.error_handler,
                            'missing_data_handler': scheme.missing_data_handler
                        }}
        }, level="member",
            context={"master_user": master_user, "member": member})


        celery_task.result_object = result_object

        celery_task.status = CeleryTask.STATUS_DONE
        celery_task.save()

    except Exception as e:

        _l.info('Exception occurred %s' % e)
        _l.info(traceback.format_exc())


@shared_task(name='integrations.complex_transaction_csv_file_import_validate', bind=True)
def complex_transaction_csv_file_import_validate(self, task_id):

    try:
        from poms.transactions.models import TransactionTypeInput
        from poms.integrations.serializers import ComplexTransactionCsvFileImport

        celery_task = CeleryTask.objects.get(pk=task_id)
        parent_celery_task = celery_task.parent

        celery_task.status = CeleryTask.STATUS_PENDING
        celery_task.save()

        master_user = celery_task.master_user
        member = celery_task.member

        instance = ComplexTransactionCsvFileImport(task_id=task_id, master_user=master_user, member=member, skip_first_line=True)

        scheme = ComplexTransactionImportScheme.objects.get(pk=celery_task.options_object['scheme_id'])

        instance.scheme = scheme
        instance.error_handling = scheme.error_handler
        instance.delimiter = scheme.delimiter
        instance.missing_data_handler = scheme.missing_data_handler
        instance.file_path = celery_task.options_object['file_path']

        _l.info('complex_transaction_csv_file_import_validate %s' % instance.file_path)

        instance.processed_rows = 0
        _l.info('complex_transaction_file_import: %s', instance)
        _l.info('complex_transaction_file_import: instance.break_on_error %s', instance.break_on_error)

        scheme_inputs = list(scheme.inputs.all())
        scheme_calculated_inputs = list(scheme.calculated_inputs.all())
        rule_scenarios = scheme.rule_scenarios.prefetch_related('transaction_type', 'fields',
                                                                'fields__transaction_type_input').all()

        _l.info('scheme %s - inputs=%s, rules=%s', scheme,
                           [(i.name, i.column) for i in scheme_inputs],
                           [(r.transaction_type.user_code) for r in rule_scenarios])

        master_user = instance.master_user
        member = instance.member

        mapping_map = {
            Account: AccountMapping,
            Currency: CurrencyMapping,
            Instrument: InstrumentMapping,
            InstrumentType: InstrumentTypeMapping,
            Counterparty: CounterpartyMapping,
            Responsible: ResponsibleMapping,
            Strategy1: Strategy1Mapping,
            Strategy2: Strategy2Mapping,
            Strategy3: Strategy3Mapping,
            DailyPricingModel: DailyPricingModelMapping,
            PaymentSizeDetail: PaymentSizeDetailMapping,
            Portfolio: PortfolioMapping,
            PriceDownloadScheme: PriceDownloadSchemeMapping,
            Periodicity: PeriodicityMapping,
            AccrualCalculationModel: AccrualCalculationModelMapping,
        }

        props_map = {
            Account: 'account',
            Currency: 'currency',
            Instrument: 'instrument',
            InstrumentType: 'instrument_type',
            Counterparty: 'counterparty',
            Responsible: 'responsible',
            Strategy1: 'strategy1',
            Strategy2: 'strategy2',
            Strategy3: 'strategy3',
            DailyPricingModel: 'daily_pricing_model',
            PaymentSizeDetail: 'payment_size_detail',
            Portfolio: 'portfolio',
            PriceDownloadScheme: 'price_download_scheme',
            Periodicity: 'periodicity',
            AccrualCalculationModel: 'accrual_calculation_model',
        }

        mapping_cache = {}

        def _get_default_relation(field):

            i = field.transaction_type_input

            model_class = i.content_type.model_class()
            model_map_class = mapping_map[model_class]

            key = props_map[model_class]

            v = None

            ecosystem_default = EcosystemDefault.objects.get(master_user=instance.master_user)

            # _l.info('key %s' % key)
            # _l.info('value %s' % value)

            if hasattr(ecosystem_default, key):
                v = getattr(ecosystem_default, key)
            else:
                v = model_map_class.objects.get(master_user=instance.master_user, value='-').content_object

            return v

        def _convert_value(field, value, error_rows):
            i = field.transaction_type_input

            if i.value_type == TransactionTypeInput.STRING:
                return str(value)

            if i.value_type == TransactionTypeInput.SELECTOR:
                return str(value)

            elif i.value_type == TransactionTypeInput.NUMBER:
                return float(value)

            elif i.value_type == TransactionTypeInput.DATE:
                if not isinstance(value, date):
                    return formula._parse_date(value)
                else:
                    return value

            elif i.value_type == TransactionTypeInput.RELATION:
                model_class = i.content_type.model_class()
                model_map_class = mapping_map[model_class]

                v = None

                try:
                    v = model_map_class.objects.get(master_user=instance.master_user, value=value).content_object
                except Exception:

                    try:

                        v = model_class.objects.get(master_user=instance.master_user, user_code=value)

                    except (model_class.DoesNotExist, KeyError):
                        v = None

                        _l.info("User code %s not found for %s " % (value, field.transaction_type_input.name))

                if not v:

                    if instance.missing_data_handler == 'set_defaults':

                        v = _get_default_relation(field)

                    else:
                        error_rows['error_message'] = error_rows[
                                                          'error_message'] + ' Can\'t find relation of ' + \
                                                      '[' + field.transaction_type_input.name + ']' + '(value:' + \
                                                      value + ')'

                return v

        def update_row_with_calculated_data(row, inputs):

            for i in scheme_calculated_inputs:

                # _l.info('update_row_with_calculated_data inputs %s' % inputs)

                try:
                    value = formula.safe_eval(i.name_expr, names=inputs)
                    row.append(value)

                except Exception:
                    _l.info('can\'t process calculated input: %s|%s', i.name, i.column, exc_info=True)
                    row.append(None)

            return row

        def _validate_process_csv_file(file, orignal_file, original_file_name):

            delimiter = instance.delimiter.encode('utf-8').decode('unicode_escape')

            reader = csv.reader(file, delimiter=delimiter, quotechar=instance.quotechar,
                                strict=False, skipinitialspace=True)

            first_row = None
            input_column_name_map = {}

            for row_index, row in enumerate(reader):

                if row_index == 0:
                    first_row = row

                    _local_index = 0
                    for item in first_row:

                        input_column_name_map[item] = _local_index
                        _local_index = _local_index + 1

                # _l.info('_validate_process_csv_file row: %s -> %s', row_index, row)
                if (row_index == 0 and instance.skip_first_line) or not row:
                    _l.info('skip first row')
                    continue

                inputs_raw = {}
                inputs = {}
                inputs_error = []
                calculated_columns_error = []

                error_rows = {
                    'level': 'info',
                    'error_message': '',
                    'inputs': inputs_raw,
                    'original_row_index': row_index,
                    'original_row': row,
                    'error_data': {
                        'columns': {
                            'imported_columns': [],
                            'calculated_columns': [],
                            'converted_imported_columns': [],
                            'transaction_type_selector': [],
                            'executed_input_expressions': []
                        },
                        'data': {
                            'imported_columns': [],
                            'calculated_columns': [],
                            'converted_imported_columns': [],
                            'transaction_type_selector': [],
                            'executed_input_expressions': []
                        }

                    },
                    'error_reaction': "Success"
                }

                for i in scheme_inputs:

                    error_rows['error_data']['columns']['imported_columns'].append(i.name)

                    if instance.scheme.column_matcher == 'index':
                        try:
                            inputs_raw[i.name] = row[i.column - 1]
                            error_rows['error_data']['data']['imported_columns'].append(row[i.column - 1])
                        except Exception:
                            _l.debug('can\'t process input: %s|%s', i.name, i.column, exc_info=True)
                            _l.debug('can\'t process inputs_raw: %s|%s', inputs_raw)
                            error_rows['error_data']['data']['imported_columns'].append(ugettext('Invalid expression'))
                            inputs_error.append(i)

                    if instance.scheme.column_matcher == 'name':

                        try:

                            _col_index = input_column_name_map[i.name]

                            inputs_raw[i.name] = row[_col_index]
                            error_rows['error_data']['data']['imported_columns'].append(row[_col_index])
                        except Exception:
                            _l.debug('can\'t process input: %s|%s', i.name, i.column, exc_info=True)
                            _l.debug('can\'t process inputs_raw: %s|%s', inputs_raw)
                            error_rows['error_data']['data']['imported_columns'].append(ugettext('Invalid expression'))
                            inputs_error.append(i)


                # _l.info('Row %s inputs_raw: %s' % (row_index, inputs_raw))

                if scheme.filter_expression:

                    # expr = Expression.parseString("a == 1 and b == 2")
                    expr = Expression.parseString(scheme.filter_expression)

                    _l.info('scheme.filter_expression %s ' % scheme.filter_expression)
                    _l.info('scheme.inputs_raw %s ' % inputs_raw)

                    if expr(inputs_raw):
                        # filter passed

                        pass
                    else:
                        _l.info("Row skipped due filter %s" % row_index)
                        continue

                for i in scheme_inputs:

                    error_rows['error_data']['columns']['converted_imported_columns'].append(
                        i.name + ': Conversion Expression ' + '(' + i.name_expr + ')')

                    try:
                        inputs[i.name] = formula.safe_eval(i.name_expr, names=inputs_raw)
                        error_rows['error_data']['data']['converted_imported_columns'].append(row[i.column - 1])
                    except Exception:
                        _l.info('can\'t process input: %s|%s', i.name, i.column, exc_info=True)
                        error_rows['error_data']['data']['converted_imported_columns'].append(
                            ugettext('Invalid expression'))
                        inputs_error.append(i)

                # _l.info('Row %s inputs_converted: %s' % (row_index, inputs))

                original_columns_count = len(row)

                row = update_row_with_calculated_data(row, inputs)

                # _l.info('Row %s inputs_with_calculated: %s' % (row_index, inputs))

                for i in scheme_calculated_inputs:

                    error_rows['error_data']['columns']['calculated_columns'].append(i.name)

                    try:

                        index = original_columns_count + i.column - 1

                        # _l.info('original_columns_count %s' % original_columns_count)
                        # _l.info('i.column %s' % i.column)
                        # _l.info('row %s' % row)

                        inputs[i.name] = row[index]

                        error_rows['error_data']['data']['calculated_columns'].append(row[index])
                    except Exception:
                        _l.info('can\'t process input: %s|%s', i.name, i.column, exc_info=True)
                        error_rows['error_data']['data']['calculated_columns'].append(ugettext('Invalid expression'))
                        calculated_columns_error.append(i)

                if inputs_error:

                    error_rows['level'] = 'error'

                    error_rows['error_message'] = error_rows['error_message'] + str(
                        ugettext('Can\'t process inputs: %(inputs)s') % {
                            'inputs': ', '.join('[' + i.name + ']' for i in inputs_error)
                        })
                    instance.error_rows.append(error_rows)

                    if instance.break_on_error:
                        error_rows['error_reaction'] = 'Break'
                        instance.error_row_index = row_index
                        instance.error_rows.append(error_rows)
                        return
                    else:
                        error_rows['error_reaction'] = 'Continue import'
                        continue

                try:
                    rule_value = formula.safe_eval(scheme.rule_expr, names=inputs)
                except Exception as e:

                    error_rows['level'] = 'error'

                    _l.info('can\'t process rule expression', exc_info=True)
                    _l.info('error %s' % e)
                    error_rows['error_message'] = error_rows['error_message'] + str(ugettext('Can\'t eval rule expression'))
                    instance.error_rows.append(error_rows)
                    if instance.break_on_error:
                        instance.error_row_index = row_index
                        error_rows['error_reaction'] = 'Break'
                        instance.error_rows.append(error_rows)
                        return
                    else:
                        error_rows['error_reaction'] = 'Continue import'
                        continue

                if not rule_value:

                    _l.info('no rule value: %s', rule_value)

                    error_rows['level'] = 'error'

                    error_rows['error_message'] = error_rows['error_message'] + str(ugettext('Rule expression is invalid'))

                    if instance.break_on_error:
                        instance.error_row_index = row_index
                        error_rows['error_reaction'] = 'Break'
                        instance.error_rows.append(error_rows)
                        return
                    else:
                        error_rows['error_reaction'] = 'Continue import'
                        continue

                # else:
                #     _l.info('rule value: %s', rule_value)

                processed_scenarios = 0
                matched_rule = False

                for scheme_rule in rule_scenarios:

                    matched_selector = False

                    selector_values = scheme_rule.selector_values.all()

                    for selector_value in selector_values:

                        if selector_value.value == rule_value:
                            matched_selector = True

                    if matched_selector:

                        processed_scenarios = processed_scenarios + 1

                        error_rows['error_data']['columns']['transaction_type_selector'].append('TType Selector')

                        try:
                            rule = scheme_rule

                            error_rows['error_data']['data']['transaction_type_selector'].append(rule_value)

                        except Exception:

                            error_rows['level'] = 'error'

                            _l.info('rule does not find: %s', rule_value, exc_info=True)
                            error_rows['error_message'] = error_rows['error_message'] + str(
                                ugettext('Can\'t find transaction type by "%(value)s"') % {
                                    'value': rule_value
                                })
                            instance.error_rows.append(error_rows)

                            error_rows['error_data']['data']['transaction_type_selector'].append(
                                ugettext('Invalid expression'))

                            if instance.break_on_error:
                                instance.error_row_index = row_index
                                error_rows['error_reaction'] = 'Break'
                                instance.error_rows.append(error_rows)
                                return
                            else:
                                error_rows['error_reaction'] = 'Continue import'
                                continue

                        _l.info('founded rule: %s -> %s', rule, rule.transaction_type)

                        fields = {}
                        fields_error = []

                        for field in rule.fields.all():

                            error_rows['error_data']['columns']['executed_input_expressions'].append(
                                field.transaction_type_input.name)

                            try:
                                field_value = formula.safe_eval(field.value_expr, names=inputs)

                                field_value = _convert_value(field, field_value, error_rows)

                                fields[field.transaction_type_input.name] = field_value

                                if hasattr(field_value, 'name'):
                                    error_rows['error_data']['data']['executed_input_expressions'].append(field_value.name)
                                else:
                                    error_rows['error_data']['data']['executed_input_expressions'].append(field_value)

                            except (Exception, ValueError, formula.InvalidExpression):

                                _l.info('can\'t process field: %s|%s', field.transaction_type_input.name,
                                                   field.transaction_type_input.pk, exc_info=True)
                                fields_error.append(field)

                                error_rows['error_data']['data']['executed_input_expressions'].append(
                                    ugettext('Invalid expression'))

                        if len(fields_error):

                            _l.info('fields (step 1): error=%s, values=%s', fields_error, fields)

                            _l.info(error_rows['error_message'])

                            inputs_messages = []

                            for field_error in fields_error:
                                message = '[' + field_error.transaction_type_input.name + '] ' + '( TType Input, TType ' + rule.transaction_type.name + ' [' + rule.transaction_type.user_code + '] )'

                                inputs_messages.append(message)

                            error_rows['error_message'] = error_rows['error_message'] + str(
                                ugettext('Can\'t process fields: %(messages)s') % {
                                    'messages': ', '.join(str(m) for m in inputs_messages)
                                })

                            error_rows['level'] = 'error'

                            if instance.break_on_error:
                                error_rows['error_reaction'] = 'Break'
                                instance.error_row_index = row_index
                                instance.error_rows.append(error_rows)
                                return
                            else:
                                error_rows['error_reaction'] = 'Continue import'
                                continue

                # print('matched_rule %s' % matched_rule)

                if processed_scenarios == 0:
                    error_rows['level'] = 'error'

                    error_rows['error_message'] = error_rows['error_message'] + str('Selector does not match')

                    if instance.break_on_error:
                        instance.error_row_index = row_index
                        error_rows['error_reaction'] = 'Break'
                        instance.error_rows.append(error_rows)
                        return
                    else:
                        error_rows['error_reaction'] = 'Continue import'

                instance.error_rows.append(error_rows)

                # if fields_error:
                #
                #     if instance.break_on_error:
                #         error_rows['error_reaction'] = 'Break'
                #         instance.error_row_index = row_index
                #         return
                #     else:
                #         error_rows['error_reaction'] = 'Continue import'
                #         continue

                instance.processed_rows = instance.processed_rows + 1
                # instance.save()

                send_websocket_message(data={
                    'type': 'transaction_import_status',
                    'payload': {
                                'parent_task_id': celery_task.parent_id,
                                'task_id': instance.task_id,
                                'state': Task.STATUS_PENDING,
                                'processed_rows': instance.processed_rows,
                                'parent_total_rows': parent_celery_task.options_object['total_rows'],
                                'total_rows': instance.total_rows,
                                'scheme_name': scheme.user_code,
                                'file_name': instance.filename}
                }, level="member",
                    context={"master_user": master_user, "member": member})

        def _row_count_xlsx(file):

            wb = load_workbook(filename=file)

            if instance.scheme.spreadsheet_active_tab_name and instance.scheme.spreadsheet_active_tab_name in wb.sheetnames:
                ws = wb[instance.scheme.spreadsheet_active_tab_name]
            else:
                ws = wb.active

            reader = []

            row_index = 0

            for r in ws.rows:
                row_index = row_index + 1

            return row_index

        def _row_count(file):

            delimiter = instance.delimiter.encode('utf-8').decode('unicode_escape')

            reader = csv.reader(file, delimiter=delimiter, quotechar=instance.quotechar,
                                strict=False, skipinitialspace=True)

            row_index = 0

            for row_index, row in enumerate(reader):
                pass
            return row_index

        instance.error_rows = []

        try:
            # with import_file_storage.open(instance.file_path, 'rb') as f:

            _l.info('Trying to open %s' % instance.file_path)
            with SFS.open(instance.file_path, 'rb') as f:
                with NamedTemporaryFile() as tmpf:
                    for chunk in f.chunks():
                        tmpf.write(chunk)
                    tmpf.flush()

                    os.link(tmpf.name, tmpf.name + '.xlsx')


                    if '.csv' in instance.file_path:

                        with open(tmpf.name, mode='rt', encoding=instance.encoding, errors='ignore') as cfr:
                            instance.total_rows = _row_count(cfr)

                        with open(tmpf.name, mode='rt', encoding=instance.encoding, errors='ignore') as cf:
                            _validate_process_csv_file(cf, f, '')

                    elif '.xlsx' in instance.file_path:

                        instance.total_rows = _row_count_xlsx(tmpf.name + '.xlsx')

                        with open(tmpf.name, mode='rt', encoding=instance.encoding, errors='ignore') as cf:
                            _validate_process_csv_file(cf, f, tmpf.name + '.xlsx')

        except Exception:
            _l.info('Can\'t process file', exc_info=True)
            instance.error_message = ugettext("Invalid file format or file already deleted.")
        finally:
            # import_file_storage.delete(instance.file_path)
            SFS.delete(instance.file_path)

        _l.info("transaction import validation completed")

        instance.error = bool(instance.error_message) or (instance.error_row_index is not None) or bool(instance.error_rows)

        # instance.stats_file_report = generate_file_report(instance, master_user, 'transaction_import.validate',
        #                                                   'Transaction Import Validation')

        send_websocket_message(data={
            'type': 'transaction_import_status',
            'payload': {
                        'parent_task_id': celery_task.parent_id,
                        'task_id': instance.task_id,
                        'state': Task.STATUS_DONE,
                        'processed_rows': instance.processed_rows,
                        'parent_total_rows': parent_celery_task.options_object['total_rows'],
                        'total_rows': instance.total_rows,
                        'file_name': instance.filename,
                        'error_rows': instance.error_rows,
                        'stats_file_report': instance.stats_file_report,
                        'scheme': scheme.id,
                        'scheme_object': {
                            'id': scheme.id,
                            'scheme_name': scheme.user_code,
                            'delimiter': scheme.delimiter,
                            'error_handler': scheme.error_handler,
                            'missing_data_handler': scheme.missing_data_handler
                        }}
        }, level="member",
            context={"master_user": master_user, "member": member})

        result_object = {
            'processed_rows': instance.processed_rows,
            'total_rows': instance.total_rows,
            'file_name': instance.filename,
            'error_rows': instance.error_rows,
            'stats_file_report': instance.stats_file_report
        }

        celery_task.result_object = result_object
        celery_task.status = CeleryTask.STATUS_DONE
        celery_task.save()

        return instance

    except Exception as e:

        _l.info('Exception occurred %s' % e)
        _l.info(traceback.format_exc())


# @shared_task(name='integrations.complex_transaction_csv_file_import_validate_parallel', bind=True)
def complex_transaction_csv_file_import_validate_parallel(task_id):

    try:

        _l.info('complex_transaction_csv_file_import_validate_parallel: task_id %s' % task_id)

        celery_task = CeleryTask.objects.get(pk=task_id)

        celery_task.status = CeleryTask.STATUS_PENDING
        celery_task.save()

        options_object = celery_task.options_object

        sub_tasks = []

        lines_per_file = 300
        header_line = None

        def _get_path(master_user, file_name, ext):
            return '%s/transaction_import_files/%s.%s' % (master_user.token, file_name, ext)

        chunk = None

        with SFS.open(celery_task.options_object['file_path'], 'rb') as f:

            _l.info("Start reading file to split it into chunks")

            ext = celery_task.options_object['file_path'].split('.')[-1]

            for lineno, line in enumerate(f):

                # _l.info('line %s' % lineno)

                if lineno == 0:
                    header_line = line
                    # _l.info('set header line %s' % lineno)

                if lineno % lines_per_file == 0:

                    if chunk is not None:
                        # _l.info("Saving chunk %s" % chunk)
                        SFS.save(chunk_path, chunk) # save working chunk before creating new one

                    chunk_filename = '%s_chunk_file_%s' % (celery_task.id, str(lineno) + '_' + str(lineno + lines_per_file))
                    chunk_path = _get_path(celery_task.master_user, chunk_filename, ext)

                    # _l.info('creating chunk file %s' % chunk_path)

                    chunk = BytesIO()
                    if lineno != 0:
                        chunk.write(header_line)

                    _l.info("creating chunk %s" % chunk_filename)

                    # _l.info('creating sub task for %s' % chunk_filename)

                    sub_task = CeleryTask.objects.create(master_user=celery_task.master_user, member=celery_task.member, parent=celery_task)

                    sub_task_options_object = copy.deepcopy(celery_task.options_object)
                    sub_task_options_object['file_path'] = chunk_path

                    sub_task.options_object = sub_task_options_object
                    sub_task.save()

                    sub_tasks.append(sub_task)

                chunk.write(line)

            _l.info("Saving last chunk")
            SFS.save(chunk_path, chunk) # save working chunk before creating new one

        _l.info('sub_tasks created %s' % len(sub_tasks))
        _l.info('original file total rows %s' % lineno)

        options_object['total_rows'] = lineno

        celery_task.options_object = options_object
        celery_task.save()

        celery_sub_tasks = []

        for sub_task in sub_tasks:

            _l.info('initializing sub_task %s' % sub_task.options_object['file_path'])

            ct = complex_transaction_csv_file_import_validate.s(task_id=sub_task.id)
            celery_sub_tasks.append(ct)

        _l.info('celery_sub_tasks len %s' % len(celery_sub_tasks))
        _l.info('celery_sub_tasks %s' % celery_sub_tasks)

        # chord(celery_sub_tasks, complex_transaction_csv_file_import_validate_parallel_finish.si(task_id=task_id)).apply_async()
        chord(celery_sub_tasks)(complex_transaction_csv_file_import_validate_parallel_finish.si(task_id=task_id))

    except Exception as e:

        _l.info('Exception occurred %s' % e)
        _l.info(traceback.format_exc())


@shared_task(name='integrations.complex_transaction_csv_file_import_by_procedure', bind=True)
def complex_transaction_csv_file_import_by_procedure(self, procedure_instance_id, transaction_file_result_id):
    with transaction.atomic():

        from poms.integrations.serializers import ComplexTransactionCsvFileImport
        from poms.procedures.models import RequestDataFileProcedureInstance

        procedure_instance = RequestDataFileProcedureInstance.objects.get(id=procedure_instance_id)
        transaction_file_result = TransactionFileResult.objects.get(id=transaction_file_result_id)

        try:

            _l.debug(
                'complex_transaction_csv_file_import_by_procedure looking for scheme %s ' % procedure_instance.procedure.scheme_user_code)

            scheme = ComplexTransactionImportScheme.objects.get(master_user=procedure_instance.master_user,
                                                                user_code=procedure_instance.procedure.scheme_user_code)

            text = "Data File Procedure %s. File is received. Decrypting file" % (
                procedure_instance.procedure.user_code)

            send_system_message(master_user=procedure_instance.master_user,
                                source="Data File Procedure Service",
                                text=text)

            _l.debug('trying to open %s' % transaction_file_result.file_path)

            with SFS.open(transaction_file_result.file_path, 'rb') as f:

                try:

                    encrypted_text = f.read()

                    rsa_cipher = RSACipher()

                    aes_key = None

                    try:
                        aes_key = rsa_cipher.decrypt(procedure_instance.private_key, procedure_instance.symmetric_key)

                        _l.debug("complex_transaction_csv_file_import_by_procedure decrypting symmetric key")

                    except Exception as e:
                        _l.debug('complex_transaction_csv_file_import_by_procedure AES Key decryption error %s' % e)

                    aes_cipher = AESCipher(aes_key)

                    decrypt_text = None

                    try:
                        decrypt_text = aes_cipher.decrypt(encrypted_text)

                        _l.debug("complex_transaction_csv_file_import_by_procedure decrypting text file")

                    except Exception as e:
                        _l.debug('complex_transaction_csv_file_import_by_procedure Text decryption error %s' % e)

                    _l.debug('complex_transaction_csv_file_import_by_procedure file decrypted')

                    _l.debug('Size of decrypted text: %s' % len(decrypt_text))


                    with NamedTemporaryFile() as tmpf:

                        _l.debug('tmpf.name %s' % tmpf.name)

                        tmpf.write(decrypt_text.encode('utf-8'))
                        tmpf.flush()

                        file_name = '%s-%s' % (timezone.now().strftime('%Y%m%d%H%M%S'), uuid.uuid4().hex)
                        file_path = '%s/data_files/%s.csv' % (procedure_instance.master_user.token, file_name)

                        SFS.save(file_path, tmpf)

                        _l.debug('complex_transaction_csv_file_import_by_procedure tmp file filled')

                        instance = ComplexTransactionCsvFileImport(scheme=scheme,
                                                                    file_path=file_path,
                                                                   missing_data_handler=scheme.missing_data_handler,
                                                                   error_handling=scheme.error_handler,
                                                                   delimiter=scheme.delimiter,
                                                                   member=procedure_instance.member,
                                                                   master_user=procedure_instance.master_user)

                    _l.debug('complex_transaction_csv_file_import_by_procedure instance: %s' % instance)

                    current_date_time = now().strftime("%Y-%m-%d-%H-%M")

                    file_name = '%s-%s' % (timezone.now().strftime('%Y%m%d%H%M%S'), uuid.uuid4().hex)
                    file_name_hash = hashlib.md5(file_name.encode('utf-8')).hexdigest()

                    file_report = FileReport()

                    file_report.upload_file(file_name='Data Procedure %s (%s).csv' % (current_date_time, file_name_hash), text=decrypt_text, master_user=procedure_instance.master_user)
                    file_report.master_user = procedure_instance.master_user
                    file_report.name = "'Transaction Import File. Procedure ' %s %s" % (procedure_instance.id, current_date_time)
                    file_report.file_name = 'Data Procedure %s (%s).csv' % (current_date_time, file_name_hash)
                    file_report.type = 'transaction_import.import'
                    file_report.notes = 'Transaction Import File. Procedure %s' % procedure_instance.id

                    file_report.save()

                    _l.debug('file_report %s' % file_report)

                    text = "Data File Procedure %s. File is received. Start Import" % (
                        procedure_instance.procedure.user_code)

                    send_system_message(master_user=procedure_instance.master_user,
                                        source="Data File Procedure Service",
                                        text=text,
                                        file_report_id=file_report.id)

                    options_object = {}
                    options_object['file_path'] = instance.file_path
                    options_object['scheme_id'] = instance.scheme.id
                    options_object['execution_context'] =  {'started_by': 'procedure'}

                    total_rows = 0

                    with SFS.open(options_object['file_path'], 'rb') as f1:

                        _l.info("Start reading file to split it into chunks")

                        for lineno, line in enumerate(f1):

                            total_rows = lineno

                    options_object['total_rows'] = total_rows

                    _l.debug('complex_transaction_csv_file_import_by_procedure total_rows %s' %  options_object['total_rows'])

                    celery_task = CeleryTask(master_user=procedure_instance.master_user,
                                             member=procedure_instance.member,
                                             options_object=options_object,
                                             type='transaction_import')

                    celery_task.save()

                    # Creating subtask
                    sub_task = CeleryTask.objects.create(master_user=celery_task.master_user, member=celery_task.member, parent=celery_task)

                    sub_task_options_object = copy.deepcopy(celery_task.options_object)

                    sub_task.options_object = sub_task_options_object

                    sub_task.save()

                    # transaction.on_commit(
                    #     lambda: complex_transaction_csv_file_import.apply_async(kwargs={'instance': instance, 'execution_context': {'started_by': 'procedure'}}))

                    # transaction.on_commit(
                    #     lambda: complex_transaction_csv_file_import.apply_async(kwargs={'task_id': sub_task.pk}))

                    celery_sub_tasks = []

                    ct = complex_transaction_csv_file_import.s(task_id=sub_task.id)
                    celery_sub_tasks.append(ct)

                    _l.info("Creating %s subtasks" % len(celery_sub_tasks))

                    # chord(celery_sub_tasks, complex_transaction_csv_file_import_parallel_finish.si(task_id=celery_task.pk)).apply_async()

                    transaction.on_commit(
                        lambda:  chord(celery_sub_tasks, complex_transaction_csv_file_import_parallel_finish.si(task_id=celery_task.pk)).apply_async())

                except Exception as e:

                    _l.debug('complex_transaction_csv_file_import_by_procedure decryption error %s' % e)

        except ComplexTransactionImportScheme.DoesNotExist:

            text = "Data File Procedure %s. Can't import file, Import scheme %s is not found" % (
            procedure_instance.procedure.user_code, procedure_instance.procedure.scheme_name)

            send_system_message(master_user=procedure_instance.master_user,
                                source="Data File Procedure Service",
                                text=text)

            _l.debug(
                'complex_transaction_csv_file_import_by_procedure scheme %s not found' % procedure_instance.procedure.scheme_name)

            procedure_instance.status = RequestDataFileProcedureInstance.STATUS_ERROR
            procedure_instance.save()


@shared_task(name='integrations.complex_transaction_csv_file_import_by_procedure_json', bind=True)
def complex_transaction_csv_file_import_by_procedure_json(self, procedure_instance_id, celery_task_id):
    with transaction.atomic():

        _l.info('complex_transaction_csv_file_import_by_procedure_json  procedure_instance_id %s celery_task_id %s' % (procedure_instance_id, celery_task_id))

        from poms.integrations.serializers import ComplexTransactionCsvFileImport
        from poms.procedures.models import RequestDataFileProcedureInstance

        procedure_instance = RequestDataFileProcedureInstance.objects.get(id=procedure_instance_id)
        celery_task = CeleryTask.objects.get(id=celery_task_id)

        try:

            _l.info(
                'complex_transaction_csv_file_import_by_procedure_json looking for scheme %s ' % procedure_instance.procedure.scheme_user_code)

            scheme = ComplexTransactionImportScheme.objects.get(master_user=procedure_instance.master_user,
                                                                user_code=procedure_instance.procedure.scheme_user_code)


            options_object = celery_task.options_object

            options_object['file_path'] = ''
            options_object['scheme_id'] = scheme.id
            options_object['execution_context'] =  {'started_by': 'procedure'}

            celery_task.options_object = options_object
            celery_task.save()

            text = "Data File Procedure %s. File is received. Importing JSON" % (
                procedure_instance.procedure.user_code)

            send_system_message(master_user=procedure_instance.master_user,
                                source="Data File Procedure Service",
                                text=text)

            ct = complex_transaction_csv_file_import.apply_async(kwargs={"task_id": celery_task.id})


        except Exception as e:

            _l.info('complex_transaction_csv_file_import_by_procedure_json e %s' % e)

            text = "Data File Procedure %s. Can't import json, Error %s" % (
                procedure_instance.procedure.user_code, e)

            send_system_message(master_user=procedure_instance.master_user,
                                source="Data File Procedure Service",
                                text=text)

            _l.debug(
                'complex_transaction_csv_file_import_by_procedure scheme %s not found' % procedure_instance.procedure.scheme_name)

            procedure_instance.status = RequestDataFileProcedureInstance.STATUS_ERROR
            procedure_instance.save()
